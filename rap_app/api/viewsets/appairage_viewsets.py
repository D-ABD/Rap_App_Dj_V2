from django.utils import timezone as dj_timezone
from django.db.models import Q, OuterRef, Subquery
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.exceptions import ValidationError, PermissionDenied
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from django.conf import settings
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import django_filters
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import extend_schema, OpenApiResponse, OpenApiTypes

from ...utils.filters import AppairageFilterSet
from ...models.appairage import Appairage, AppairageActivite, AppairageStatut
from ...models.commentaires_appairage import CommentaireAppairage
from ..serializers.appairage_serializers import (
    AppairageSerializer,
    AppairageListSerializer,
    AppairageCreateUpdateSerializer,
    AppairageMetaSerializer,
    CommentaireAppairageSerializer,
)
from ..permissions import IsStaffOrAbove, is_staff_or_staffread
from ..paginations import RapAppPagination


# ==========================================================================
# APPARIAGE VIEWSET
# ==========================================================================
class AppairageViewSet(viewsets.ModelViewSet):
    base_queryset = (
        Appairage.objects.all()
        .select_related(
            "candidat",
            "candidat__formation",
            "candidat__formation__centre",
            "candidat__formation__type_offre",
            "candidat__formation__statut",
            "partenaire",
            "formation",
            "formation__centre",
            "formation__type_offre",
            "formation__statut",
            "created_by",
            "updated_by",
        )
        .prefetch_related("historiques", "commentaires")
    )

    permission_classes = [IsStaffOrAbove]
    pagination_class = RapAppPagination
    filterset_class = AppairageFilterSet
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["statut", "formation", "candidat", "partenaire", "created_by"]

    search_fields = [
        "candidat__nom",
        "candidat__prenom",
        "partenaire__nom",
        "partenaire_contact_nom", 
        "formation__nom",
        "candidat__formation__nom",
        "formation__centre__nom",
        "candidat__formation__centre__nom",
        "commentaires__body",
        "retour_partenaire",
        "created_by__first_name",
        "created_by__last_name",
        "created_by__username",
        "created_by__email",
    ]

    ordering_fields = [
        "date_appairage",
        "statut",
        "formation__nom",
        "formation__centre__nom",
        "updated_at",
        "created_at",
    ]

    # ------------------------- helpers scope/permission -------------------------
    def _is_admin_like(self, user) -> bool:
        return getattr(user, "is_superuser", False) or (
            hasattr(user, "is_admin") and user.is_admin()
        )

    def _staff_centre_ids(self, user):
        if self._is_admin_like(user):
            return None
        if is_staff_or_staffread(user):
            return list(user.centres.values_list("id", flat=True))
        return []

    def _scope_qs_to_user_centres(self, qs):
        user = self.request.user
        if not user.is_authenticated:
            return qs.none()

        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            return qs.none()

        centre_ids = self._staff_centre_ids(user)
        if centre_ids is None:  # admin/superuser
            return qs

        if centre_ids:
            return qs.filter(
                Q(formation__centre_id__in=centre_ids)
                | Q(candidat__formation__centre_id__in=centre_ids)
            ).distinct()

        return qs.none()

    def _assert_staff_can_use_formation(self, formation):
        if not formation:
            return
        user = self.request.user
        if self._is_admin_like(user):
            return
        if is_staff_or_staffread(user):
            allowed = set(user.centres.values_list("id", flat=True))
            if getattr(formation, "centre_id", None) not in allowed:
                raise PermissionDenied("Formation hors de votre périmètre (centre).")

    # ------------------------- helpers export -------------------------
    def _user_display(self, user):
        if not user:
            return ""
        return f"{user.get_full_name()} ({user.username})" if hasattr(user, "username") else str(user)

    def _partenaire_email(self, a):
        return getattr(a.partenaire, "email", "") or getattr(a, "partenaire_email", "") or ""

    def _partenaire_telephone(self, a):
        return getattr(a.partenaire, "telephone", "") or getattr(a, "partenaire_telephone", "") or ""

    def _formation_id(self, a):
        return getattr(a.formation, "id", "") or ""

    def _formation_label(self, a):
        return getattr(a.formation, "nom", "") or ""

    def _formation_type_offre(self, a):
        return getattr(getattr(a.formation, "type_offre", None), "libelle", "") or getattr(a, "formation_type_offre", "")

    def _formation_num_offre(self, a):
        return getattr(a.formation, "num_offre", "") or getattr(a, "formation_numero_offre", "")

    # ------------------------------- DRF hooks ---------------------------------
    def get_serializer_class(self):
        if self.action == "list":
            return AppairageListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return AppairageCreateUpdateSerializer
        return AppairageSerializer


    def perform_create(self, serializer):
        user = self.request.user
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            raise PermissionDenied("Les candidats/stagiaires ne peuvent pas créer d’appairage.")

        formation_payload = serializer.validated_data.get("formation")
        candidat_payload = serializer.validated_data.get("candidat")
        formation = (
            formation_payload
            or (
                getattr(getattr(candidat_payload, "formation", None), "pk", None)
                and candidat_payload.formation
            )
            or None
        )

        if formation:
            self._assert_staff_can_use_formation(formation)

        partenaire_payload = serializer.validated_data.get("partenaire")
        if Appairage.objects.filter(
            candidat=candidat_payload, partenaire=partenaire_payload, formation=formation
        ).exists():
            raise ValidationError(
                {"detail": "Un appairage existe déjà pour ce candidat, ce partenaire et cette formation."}
            )

        instance = serializer.save(
            created_by=user, formation=formation or serializer.validated_data.get("formation")
        )
        if hasattr(instance, "set_user"):
            instance.set_user(user)
        try:
            instance.save(user=user)
        except TypeError:
            instance.save()

    def perform_update(self, serializer):
        user = self.request.user
        instance = serializer.instance
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            raise PermissionDenied("Les candidats/stagiaires ne peuvent pas modifier un appairage.")

        data_formation = serializer.validated_data.get("formation", instance.formation)
        if data_formation:
            self._assert_staff_can_use_formation(data_formation)

        instance = serializer.save(formation=data_formation)
        if hasattr(instance, "set_user"):
            instance.set_user(user)
        try:
            instance.save(user=user)
        except TypeError:
            instance.save()

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        serializer = AppairageMetaSerializer(instance={}, context={"request": request})
        return Response(serializer.data)

    # ---------------------------- Commentaires ----------------------------
    @action(detail=True, methods=["get", "post"], url_path="commentaires")
    def commentaires(self, request, pk=None):
        appairage = self.get_object()
        if request.method == "GET":
            qs = appairage.commentaires.order_by("-created_at")
            serializer = CommentaireAppairageSerializer(qs, many=True)
            return Response(serializer.data)

        if request.method == "POST":
            serializer = CommentaireAppairageSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(created_by=request.user, appairage=appairage)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        

    def get_queryset(self):
        qs = self.base_queryset

        # 🗒️ Annoter le dernier commentaire
        last_comment_qs = (
            CommentaireAppairage.objects.filter(appairage=OuterRef("pk"))
            .order_by("-created_at")
            .values("body")[:1]
        )
        qs = qs.annotate(last_commentaire=Subquery(last_comment_qs))

        # 🧩 Filtrage explicite par activité
        activite = self.request.query_params.get("activite")
        if activite in [AppairageActivite.ACTIF, AppairageActivite.ARCHIVE]:
            # Si un type d'activité est spécifié → filtrage strict
            qs = qs.filter(activite=activite)
        else:
            # ⚙️ Par défaut, on exclut les archivés sauf si `avec_archivees=true`
            avec_archivees = self.request.query_params.get("avec_archivees")
            if not (avec_archivees and str(avec_archivees).lower() in ["1", "true", "yes", "on"]):
                qs = qs.exclude(activite=AppairageActivite.ARCHIVE)

        # 🔒 Restriction par centre selon l’utilisateur connecté
        return self._scope_qs_to_user_centres(qs)

    # ✅ Permet d’afficher les appairages archivés en détail
    def retrieve(self, request, *args, **kwargs):
        # On va chercher dans base_queryset pour ne pas exclure les archivés
        obj = get_object_or_404(self.base_queryset, pk=kwargs.get("pk"))
        serializer = self.get_serializer(obj)
        return Response(serializer.data)

    # ----------------------------------------------------------------------
    # 📦 Archivage
    # ----------------------------------------------------------------------

    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        """
        Archive un appairage : il ne sera plus visible par défaut.
        """
        # 🔁 utiliser base_queryset pour inclure tous les appairages
        appairage = get_object_or_404(self.base_queryset, pk=pk)

        if appairage.activite == AppairageActivite.ARCHIVE:
            return Response({"detail": "Déjà archivé."}, status=status.HTTP_400_BAD_REQUEST)

        if hasattr(appairage, "archiver"):
            appairage.archiver(user=request.user)
        else:
            appairage.activite = AppairageActivite.ARCHIVE
            appairage.save(user=request.user, update_fields=["activite"])

        return Response({"status": "archived"}, status=status.HTTP_200_OK)


    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """
        Désarchive un appairage : il redevient actif.
        """
        # 🔁 idem : on prend base_queryset pour inclure aussi les archivés
        appairage = get_object_or_404(self.base_queryset, pk=pk)

        if appairage.activite != AppairageActivite.ARCHIVE:
            return Response({"detail": "Cet appairage n’est pas archivé."}, status=status.HTTP_400_BAD_REQUEST)

        if hasattr(appairage, "desarchiver"):
            appairage.desarchiver(user=request.user)
        else:
            appairage.activite = AppairageActivite.ACTIF
            appairage.save(user=request.user, update_fields=["activite"])

        return Response({"status": "unarchived"}, status=status.HTTP_200_OK)

    # ----------------------------------------------------------------------
    # 🧮 Export
    # ----------------------------------------------------------------------
    def _get_export_queryset(self, request):
        qs = self.filter_queryset(self.get_queryset())

        # ⚙️ Exclure les archivés sauf si demandé
        avec_archivees = request.data.get("avec_archivees") or request.query_params.get("avec_archivees")
        if not (avec_archivees and str(avec_archivees).lower() in ["1", "true", "yes", "on"]):
            qs = qs.exclude(activite=AppairageActivite.ARCHIVE)

        # IDs filtrés
        ids = request.data.get("ids") or request.query_params.get("ids")
        if ids:
            if isinstance(ids, str):
                ids = [int(x) for x in ids.split(",") if x.isdigit()]
            elif isinstance(ids, list):
                ids = [int(x) for x in ids if str(x).isdigit()]
            qs = qs.filter(id__in=ids)

        return qs.select_related(
            "candidat",
            "candidat__formation",
            "candidat__formation__centre",
            "formation",
            "formation__centre",
            "partenaire",
            "created_by",
            "updated_by",
        ).prefetch_related("commentaires")


    @extend_schema(
        summary="Exporter les appairages (Excel)",
        description="Exporte les appairages filtrés au format Excel (.xlsx).",
        responses={
            200: OpenApiResponse(
                description="Fichier Excel généré avec succès.",
                response=OpenApiTypes.BINARY,
                examples=None,
            )
        },
    )
    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self._get_export_queryset(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Appairages"

        # ==========================================================
        # 🖼️ Logo Rap_App
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre et date
        # ==========================================================
        ws.merge_cells("B1:Z1")
        ws["B1"] = "Export des appairages — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:Z2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "Activité (code)", "Date appairage", "Statut (code)",
            "Candidat",
            "Partenaire", "Contact", "Email", "Téléphone",
            "Formation", "Centre",
            "Type d’offre", "N° Offre", "Statut formation",
            "Places totales", "Places disponibles",
            "Date début", "Date fin",
            "Retour partenaire", "Date retour",
            "Créé par (nom)", "Créé le",
            "Maj par (nom)", "Maj le",
            "Dernier commentaire", "Commentaires",
        ]
        ws.append(headers)

        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[header_row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        even_fill = PatternFill("solid", fgColor="F7FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        def _safe(obj, path, default=""):
            try:
                for part in path.split("."):
                    obj = getattr(obj, part)
                    if obj is None:
                        return default
                return obj
            except Exception:
                return default

        def _to_text(v):
            if v is None:
                return ""
            if callable(v):
                try:
                    v = v()
                except Exception:
                    v = str(v)
            if isinstance(v, (int, float, str, bool)):
                return str(v)
            if isinstance(v, (list, tuple, set)):
                return "\n".join(map(_to_text, v))
            if isinstance(v, dict):
                return "; ".join(f"{k}: {v}" for k, v in v.items())
            if hasattr(v, "strftime"):
                try:
                    return v.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    return str(v)
            if hasattr(v, "nom_complet"):
                return v.nom_complet
            if hasattr(v, "nom"):
                return v.nom
            return str(v)

        def _compute_places_disponibles(a):
            f = getattr(a, "formation", None)
            if not f:
                return ""
            inscrits_total = (getattr(f, "inscrits_crif", 0) or 0) + (getattr(f, "inscrits_mp", 0) or 0)
            prevus_total = (getattr(f, "prevus_crif", 0) or 0) + (getattr(f, "prevus_mp", 0) or 0)
            cap = getattr(f, "cap", None)
            if cap is not None:
                return max(int(cap) - int(inscrits_total), 0)
            if prevus_total:
                return max(int(prevus_total) - int(inscrits_total), 0)
            return ""

        for i, a in enumerate(qs, start=1):
            commentaires_text = ""
            try:
                if hasattr(a, "commentaires"):
                    coms = a.commentaires.all().order_by("-created_at")
                    commentaires_text = "\n".join(
                        f"- {getattr(c.created_by, 'get_full_name', lambda: str(c.created_by))()}: {getattr(c, 'body', '')}"
                        for c in coms
                    )
            except Exception:
                commentaires_text = ""

            row = [
                _to_text(_safe(a, "activite")),
                a.date_appairage.strftime("%d/%m/%Y") if _safe(a, "date_appairage") else "",
                _to_text(_safe(a, "statut")),
                _to_text(_safe(a, "candidat.nom_complet")) or _to_text(_safe(a, "candidat")),
                _to_text(_safe(a, "partenaire.nom")),
                _to_text(_safe(a, "partenaire.contact_nom")),
                _to_text(_safe(a, "partenaire.contact_email")),
                _to_text(_safe(a, "partenaire.contact_telephone")),
                _to_text(_safe(a, "formation.nom")),
                _to_text(_safe(a, "formation.centre.nom")),
                _to_text(_safe(a, "formation.type_offre.nom")),
                _to_text(_safe(a, "formation.num_offre")),
                _to_text(_safe(a, "formation.statut")),
                _to_text(_safe(a, "formation.cap")),
                _to_text(_compute_places_disponibles(a)),
                _to_text(_safe(a, "formation.start_date")),
                _to_text(_safe(a, "formation.end_date")),
                _to_text(getattr(a, "retour_partenaire", "")),
                a.date_retour.strftime("%d/%m/%Y") if _safe(a, "date_retour") else "",
                _to_text(_safe(a, "created_by.get_full_name")),
                a.created_at.strftime("%d/%m/%Y %H:%M") if _safe(a, "created_at") else "",
                _to_text(_safe(a, "updated_by.get_full_name")),
                a.updated_at.strftime("%d/%m/%Y %H:%M") if _safe(a, "updated_at") else "",
                _to_text(getattr(a, "last_commentaire", "")),
                commentaires_text,
            ]
            ws.append(row)

            fill = even_fill if i % 2 == 0 else odd_fill
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrapText=True)
                val = str(cell.value).strip().lower() if cell.value else ""

                # 🎨 Couleur dynamique selon les colonnes
                if j == 3:  # Statut
                    if "actif" in val:
                        cell.font = Font(color="008000", bold=True)
                    elif "inactif" in val or "refus" in val:
                        cell.font = Font(color="C00000", bold=True)
                    elif "archive" in val:
                        cell.font = Font(color="7F7F7F", italic=True)
                    else:
                        cell.font = Font(color="1F4E78")
                elif j == 14:  # Places totales → bleu fixe
                    cell.font = Font(color="1F4E78", bold=True)
                elif j == 15:  # Places disponibles → dynamique
                    try:
                        num = int(float(val.replace(",", ".").strip()))
                    except Exception:
                        num = None
                    if num is None:
                        cell.font = Font(color="000000")
                    elif num == 0:
                        cell.font = Font(color="006100", bold=True)  # 🟩 complet
                    elif num <= 4:
                        cell.font = Font(color="E46C0A", bold=True)  # 🟧 presque plein
                    elif num <= 9:
                        cell.font = Font(color="1F4E78", bold=True)  # 🟦 disponible
                    else:
                        cell.font = Font(color="C00000", bold=True)  # 🟥 très disponible
                elif j in [18, 19]:
                    cell.font = Font(color="548235")
                elif j in [20, 22]:
                    cell.font = Font(color="7030A0")
                else:
                    cell.font = Font(color="000000")
            ws.row_dimensions[ws.max_row].height = 26

        # ==========================================================
        # 📊 Filtres + gel de l’en-tête
        # ==========================================================
        end_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        # ==========================================================
        # 📏 Largeurs auto-ajustées
        # ==========================================================
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            column_letter = get_column_letter(col_cells[0].column)
            adjusted_width = min(length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Pied de page
        ws.oddFooter.center.text = f"© Rap_App — export du {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        # ==========================================================
        # 📤 Génération du fichier
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        binary = buffer.getvalue()
        buffer.close()

        filename = f'appairages_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary)
        return response
        