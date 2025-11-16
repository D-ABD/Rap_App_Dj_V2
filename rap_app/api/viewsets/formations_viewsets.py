import csv
import logging
from django.http import HttpResponse
from django.db.models import Q, Count
from django.db.models.functions import TruncMonth
from io import BytesIO
from pathlib import Path
import datetime
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone as dj_timezone
from django.templatetags.static import static
from django.conf import settings
from drf_spectacular.utils import extend_schema
import pytz
from rest_framework.decorators import action
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from rest_framework import serializers
from django.db import transaction

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
from django.utils import timezone as dj_timezone
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils import timezone


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.views import APIView



from ..roles import is_admin_like, is_staff_or_staffread, staff_centre_ids

from ...models.statut import Statut
from ...models.types_offre import TypeOffre
from ...models.formations import Formation
from ...api.paginations import RapAppPagination
from ...api.permissions import IsStaffOrAbove, UserVisibilityScopeMixin
from ...api.serializers.formations_serializers import (
    FormationCreateSerializer,
    FormationListSerializer,
    FormationDetailSerializer,
)

logger = logging.getLogger("application.api")

from bs4 import BeautifulSoup
import re

def strip_html_tags_pretty(html: str) -> str:
    """
    Supprime les balises HTML et conserve un format lisible (sauts de ligne, indentation légère).
    """
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")

    # Ajoute un retour à la ligne avant certains blocs
    for tag in soup.find_all(["p", "li", "div", "br"]):
        tag.insert_before("\n")

    text = soup.get_text(separator=" ", strip=True)
    # Nettoie les espaces et retours multiples
    text = re.sub(r"\s*\n\s*", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

# 🔎 Ajoute ce mini backend pour utiliser ?texte= comme paramètre de recherche
class FormationSearchFilter(filters.SearchFilter):
    search_param = "texte"   # ex: ?texte=Responsable
    
@extend_schema(tags=["Formations"])
class FormationViewSet(UserVisibilityScopeMixin, viewsets.ModelViewSet):
    """
    📚 ViewSet pour gérer les formations.
    Accès :
      - admin/superadmin → accès complet
      - staff/staff_read → accès limité à leurs centres
      - autres utilisateurs → uniquement leurs propres objets (via UserVisibilityScopeMixin)
    """
    queryset = Formation.objects.all()
    permission_classes = [IsStaffOrAbove]
    pagination_class = RapAppPagination

    # ⬇️ remplace SearchFilter par FormationSearchFilter (texte)
    filter_backends = [
            DjangoFilterBackend,
            FormationSearchFilter,     # accepte ?texte=...
            filters.SearchFilter,  # accepte ?search=...
            filters.OrderingFilter
        ]
    filterset_fields = ["centre", "type_offre", "statut", "created_by", "start_date"]
    serializer_class = FormationListSerializer

    # ✅ DRF cherchera dans ces champs quand ?texte= est présent
    search_fields = ["nom", "num_offre", "centre__nom", "type_offre__nom","assistante",]
    ordering_fields = ["start_date", "end_date", "nom", "centre__nom",  "created_at"]
    # 👇 Ordre par défaut : formations les plus proches d'abord
    ordering = ["start_date"]
    
    # ---------- Helpers FK ----------
    def _normalize_payload_for_fk(self, data):
        """
        Accepte à la fois:
          - *_id:  { "type_offre_id": 2, "statut_id": 3, "centre_id": 1 }
          - nested: { "type_offre": {"id":2}, "statut": {"id":3}, "centre": {"id":1} }
        → Retourne un dict propre avec *_id normalisés.
        """
        p = dict(data)  # typiquement QueryDict → on copie
        # si nested, on extrait l'id
        for field in ("centre", "type_offre", "statut"):
            obj = p.get(field)
            if isinstance(obj, dict) and "id" in obj:
                p[f"{field}_id"] = obj.get("id")
        # cast string → int si besoin
        for fk in ("centre_id", "type_offre_id", "statut_id"):
            if fk in p and p[fk] in ("", None):
                p[fk] = None
            elif fk in p:
                try:
                    p[fk] = int(p[fk])
                except (TypeError, ValueError):
                    pass
        return p

    def _ensure_required_refs(self, payload):
        missing = []
        if not payload.get("centre_id"):
            missing.append("centre_id (ou centre.id)")
        if not payload.get("type_offre_id"):
            missing.append("type_offre_id (ou type_offre.id)")
        if not payload.get("statut_id"):
            missing.append("statut_id (ou statut.id)")
        if missing:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({"detail": f"Champs obligatoires manquants: {', '.join(missing)}"})

    # ---------- Serializer par action ----------
    def get_serializer_class(self):
        if self.action == "list":
            return FormationListSerializer
        if self.action == "retrieve":
            return FormationDetailSerializer
        if self.action == "create":
            # ⬅️ input minimal (exige type_offre + statut)
            return FormationCreateSerializer
        if self.action in ["update", "partial_update"]:
            # ⬅️ input large : on autorise la MAJ de TOUS les champs
            return FormationDetailSerializer
        return super().get_serializer_class()

    # ---------- CREATE ----------
    @extend_schema(
        summary="Créer une formation",
        request=FormationCreateSerializer,           # ✅ cohérent avec get_serializer_class
        responses={201: FormationDetailSerializer}
    )
    def create(self, request, *args, **kwargs):
        payload = self._normalize_payload_for_fk(request.data)
        self._ensure_required_refs(payload)          # ✅ type_offre + statut exigés

        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            formation = serializer.save()

        # Rechargement avec relations pour une réponse complète
        formation = (
            Formation.objects
            .select_related("centre", "type_offre", "statut")
            .get(pk=formation.pk)
        )
        response_serializer = FormationDetailSerializer(formation, context={"request": request})

        return Response(
            {"success": True, "message": "Formation créée avec succès.", "data": response_serializer.data},
            status=status.HTTP_201_CREATED,
        )

    # ---------- UPDATE (tous champs) ----------
    @extend_schema(
        summary="Mettre à jour une formation",
        request=FormationDetailSerializer,           # ✅ accepte tous les champs en entrée
        responses={200: FormationDetailSerializer}
    )
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        payload = self._normalize_payload_for_fk(request.data)  # accepte nested ou *_id

        serializer = self.get_serializer(instance, data=payload, partial=True)  # PATCH-like
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            formation = serializer.save()

        formation = (
            Formation.objects
            .select_related("centre", "type_offre", "statut")
            .get(pk=formation.pk)
        )
        response_serializer = FormationDetailSerializer(formation, context={"request": request})
        return Response(
            {"success": True, "message": "Formation mise à jour avec succès.", "data": response_serializer.data},
            status=status.HTTP_200_OK,
        )


    # ---------- Scope centres ----------
    def _restrict_to_user_centres(self, qs):
        u = self.request.user

        # 🔑 Admins / superadmins → accès complet
        if is_admin_like(u):
            return qs

        # 👩‍💼 Staff ou StaffRead → limité à leurs centres
        if is_staff_or_staffread(u):
            centres = staff_centre_ids(u)
            (f"👤 {u.username} ({u.role}) → centres visibles: {centres}")
            if not centres:
                return qs.none()
            return qs.filter(centre_id__in=centres)

        # 🚫 Autres → `UserVisibilityScopeMixin` va gérer created_by
        return qs

    def get_queryset(self):
        """
        Retourne le queryset filtré selon le périmètre utilisateur et les paramètres :
        - activite: active / archivee / en_cours / terminee / annulee
        - dans: 4w / 3m / 6m (période à venir)
        - avec_archivees: inclut les archivées si demandé
        """
        qs = Formation.objects.all_including_archived()
        qs = self._restrict_to_user_centres(qs)

        params = self.request.query_params
        activite = params.get("activite")
        dans = params.get("dans")
        now = timezone.now().date()
        annee = params.get("annee")   # <--- OK ici
        if annee:
            try:
                annee = int(annee)
                qs = qs.filter(
                    Q(start_date__year=annee)
                    | Q(end_date__year=annee)
                )
            except ValueError:
                pass  # année invalide → on ignore proprement

        # 🔹 Filtrage par activité
        if activite:
            if activite in ["active", "archivee"]:
                qs = qs.filter(activite=activite)
            elif activite == "en_cours":
                qs = qs.filter(start_date__lte=now, end_date__gte=now)
            elif activite == "terminee":
                qs = qs.filter(end_date__lt=now)
            elif activite == "annulee":
                qs = qs.filter(statut__nom__icontains="annul")
        else:
            # Par défaut, exclut les archivées sauf si demandé
            avec_archivees = params.get("avec_archivees")
            if not (avec_archivees and str(avec_archivees).lower() in ["1", "true", "yes", "on"]):
                qs = qs.exclude(activite="archivee")

        # 🔹 Filtrage par période "à venir"
        if dans:
            try:
                if dans == "4w":
                    limite = now + datetime.timedelta(weeks=4)
                elif dans == "3m":
                    limite = now + datetime.timedelta(days=90)
                elif dans == "6m":
                    limite = now + datetime.timedelta(days=180)
                elif dans.isdigit():
                    limite = now + datetime.timedelta(days=int(dans))
                else:
                    limite = None

                if limite:
                    qs = qs.filter(start_date__gte=now, start_date__lte=limite)
            except Exception as e:
                logger.warning(f"[filtres formations] paramètre 'dans' invalide ({dans}) : {e}")

        return qs



    def get_object(self):
        """
        🔓 Permet d'accéder aussi aux formations archivées.
        """
        pk = self.kwargs.get(self.lookup_field, None)
        qs = Formation.objects.all_including_archived().select_related("centre", "type_offre", "statut")
        qs = self._restrict_to_user_centres(qs)
        return qs.get(pk=pk)

    @extend_schema(
        summary="Lister les formations",
        description="Retourne une liste paginée des formations avec filtres disponibles.",
        parameters=[
            OpenApiParameter("texte", str, description="Recherche texte (nom, numéro d’offre, centre, type d’offre)"),
            OpenApiParameter("type_offre", str, description="ID du type d'offre"),
            OpenApiParameter("centre", str, description="ID du centre"),
            OpenApiParameter("statut", str, description="ID du statut"),
            OpenApiParameter("date_debut", str, description="Date de début minimale (AAAA-MM-JJ)"),
            OpenApiParameter("date_fin", str, description="Date de fin maximale (AAAA-MM-JJ)"),
            OpenApiParameter("places_disponibles", str, description="Filtre les formations avec des places disponibles"),
            OpenApiParameter("tri", str, description="Alias de tri (équivalent à ?ordering=, ex: -start_date, nom)"),
        ],
        responses={200: OpenApiResponse(response=FormationListSerializer(many=True))}
    )
    def list(self, request, *args, **kwargs):
        params = request.query_params

        # ✅ Applique: DjangoFilterBackend + SearchFilter(texte) + OrderingFilter(ordering)
        qs = self.filter_queryset(self.get_queryset())
        
        # ⏳ Filtre temporel : ?dans=4w / 3m / 6m / N (jours)
        dans = params.get("dans")
        if dans:
            try:
                qs = Formation.objects.formations_a_venir(dans=dans)
            except Exception as e:
                logger.warning(f"Filtre 'dans' ignoré (paramètre invalide '{dans}') : {e}")

        # ⬇️ Compléments non gérés par DjangoFilterBackend
        if params.get("date_debut"):
            qs = qs.filter(date_debut__date__gte=params.get("date_debut"))
        if params.get("date_fin"):
            qs = qs.filter(date_fin__date__lte=params.get("date_fin"))
        if params.get("places_disponibles") == "true":
            qs = qs.filter(places_disponibles__gt=0)

        # 🔁 Alias rétro-compat: ?tri=... (en plus de ?ordering=... déjà géré par OrderingFilter)
        tri = params.get("tri")
        if tri:
            try:
                qs = qs.order_by(tri)
            except Exception as e:
                logger.warning(f"Tri ignoré (paramètre invalide '{tri}') : {e}")

        # 📄 Pagination + sérialisation (une seule fois, après tous les filtres)
        page = self.paginate_queryset(qs)
        serializer = self.get_serializer(page or qs, many=True)

        if page is not None:
            return Response({
                "success": True,
                "message": "Liste paginée des formations",
                "data": {
                    "count": self.paginator.page.paginator.count,
                    "results": serializer.data,
                },
            })

        return Response({
            "success": True,
            "message": "Liste complète des formations",
            "data": {
                "count": len(serializer.data),
                "results": serializer.data,
            },
        })


    # ---------- Actions annexes (toutes restreintes aussi) ----------

    @extend_schema(summary="Filtres disponibles (centres, statuts, types d’offre, activités, périodes à venir)")
    @action(detail=False, methods=["get"])
    def filtres(self, request):
        user = request.user
        ref_complet = str(request.query_params.get("ref_complet", "")).lower() in {"1", "true", "yes", "on"}

        qs = self.get_queryset()

        # ✅ Centres : si admin/superadmin → tous, sinon selon périmètre
        if is_admin_like(user):
            from ...models.centres import Centre  # ⚠️ adapte l’import selon ton projet
            centres_qs = (
                Centre.objects.filter(is_active=True)
                .values_list("id", "nom")
                .order_by("nom")
            )
        else:
            centres_qs = (
                qs.values_list("centre_id", "centre__nom")
                .distinct()
                .order_by("centre__nom")
            )

        centres = [{"id": c[0], "nom": c[1]} for c in centres_qs if c[0]]

        # ✅ Statuts et types d’offre
        if ref_complet:
            statuts_qs = Statut.objects.all().values_list("id", "nom").order_by("nom")
            type_offres_qs = TypeOffre.objects.all().values_list("id", "nom").order_by("nom")
            statuts = [{"id": s[0], "nom": s[1]} for s in statuts_qs]
            type_offres = [{"id": t[0], "nom": t[1]} for t in type_offres_qs]
        else:
            statuts_qs = (
                qs.values_list("statut_id", "statut__nom")
                .distinct()
                .order_by("statut__nom")
            )
            type_offres_qs = (
                qs.values_list("type_offre_id", "type_offre__nom")
                .distinct()
                .order_by("type_offre__nom")
            )
            statuts = [{"id": s[0], "nom": s[1]} for s in statuts_qs if s[0]]
            type_offres = [{"id": t[0], "nom": t[1]} for t in type_offres_qs if t[0]]

        # ✅ Périodes temporelles "à venir"
        periodes_a_venir = [
            {"code": "4w", "libelle": "Dans les 4 semaines"},
            {"code": "3m", "libelle": "Dans les 3 mois"},
            {"code": "6m", "libelle": "Dans les 6 mois"},
            {"code": "180", "libelle": "Dans les 6 mois (approximatif)"},
        ]

        # ✅ Activités enrichies (pour filtres front)
        activites = [
            {"code": "active", "libelle": "Active"},
            {"code": "en_cours", "libelle": "En cours"},
            {"code": "terminee", "libelle": "Terminée"},
            {"code": "annulee", "libelle": "Annulée"},
            {"code": "archivee", "libelle": "Archivée"},
        ]

        return Response(
            {
                "success": True,
                "data": {
                    "centres": centres,
                    "statuts": statuts,
                    "type_offres": type_offres,
                    "activites": activites,
                    "periodes_a_venir": periodes_a_venir,
                },
            }
        )



    @extend_schema(summary="Obtenir l'historique d'une formation")
    @action(detail=True, methods=["get"])
    def historique(self, request, pk=None):
        data = [h.to_serializable_dict() for h in self.get_object().get_historique()]
        return Response({"success": True, "data": data})

    @extend_schema(summary="Lister les partenaires d'une formation")
    @action(detail=True, methods=["get"])
    def partenaires(self, request, pk=None):
        data = [p.to_serializable_dict() for p in self.get_object().get_partenaires()]
        return Response({"success": True, "data": data})

    @extend_schema(summary="Lister les commentaires d'une formation")
    @action(detail=True, methods=["get"])
    def commentaires(self, request, pk=None):
        f = self.get_object()
        limit = request.query_params.get("limit")
        with_saturation = request.query_params.get("saturation") == "true"
        qs = f.get_commentaires(include_saturation=with_saturation, limit=int(limit) if limit else None)
        return Response({
            "success": True,
            "data": [c.to_serializable_dict(include_full_content=True) for c in qs]
        })

    @extend_schema(summary="Lister les documents d'une formation")
    @action(detail=True, methods=["get"])
    def documents(self, request, pk=None):
        est_public = request.query_params.get("est_public")
        est_public = est_public.lower() == "true" if est_public is not None else None
        docs = self.get_object().get_documents(est_public)
        return Response({"success": True, "data": [d.to_serializable_dict() for d in docs]})

    @extend_schema(summary="Lister les prospections liées à une formation")
    @action(detail=True, methods=["get"])
    def prospections(self, request, pk=None):
        formation = self.get_object()
        # ✅ Correction : utilise le related_name défini dans Prospection.formation
        prosps = formation.prospections.all()
        return Response({
            "success": True,
            "data": [p.to_serializable_dict() for p in prosps]
        })

    @extend_schema(summary="Ajouter un commentaire à une formation")
    @action(detail=True, methods=["post"])
    def ajouter_commentaire(self, request, pk=None):
        try:
            c = self.get_object().add_commentaire(
                user=request.user,
                contenu=request.data.get("contenu"),
                saturation=request.data.get("saturation")
            )
            return Response({"success": True, "data": c.to_serializable_dict()})
        except Exception as e:
            logger.exception("Ajout commentaire échoué")
            return Response({"success": False, "message": str(e)}, status=400)

    @extend_schema(summary="Ajouter un événement à une formation")
    @action(detail=True, methods=["post"])
    def ajouter_evenement(self, request, pk=None):
        try:
            e = self.get_object().add_evenement(
                type_evenement=request.data.get("type_evenement"),
                event_date=request.data.get("event_date"),
                details=request.data.get("details"),
                description_autre=request.data.get("description_autre"),
                user=request.user
            )
            return Response({"success": True, "data": e.to_serializable_dict()})
        except Exception as e:
            logger.exception("Ajout événement échoué")
            return Response({"success": False, "message": str(e)}, status=400)

    @extend_schema(summary="Ajouter un document à une formation")
    @action(detail=True, methods=["post"])
    def ajouter_document(self, request, pk=None):
        try:
            doc = self.get_object().add_document(
                user=request.user,
                fichier=request.FILES.get("fichier"),
                nom_fichier=request.data.get("nom_fichier"),
                type_document=request.data.get("type_document")
            )
            return Response({"success": True, "data": doc.to_serializable_dict()})
        except Exception as e:
            logger.exception("Ajout document échoué")
            return Response({"success": False, "message": str(e)}, status=400)

    @extend_schema(summary="Dupliquer une formation")
    @action(detail=True, methods=["post"])
    def dupliquer(self, request, pk=None):
        try:
            f = self.get_object().duplicate(user=request.user)
            return Response({"success": True, "message": "Formation dupliquée", "data": f.to_serializable_dict()})
        except Exception as e:
            logger.exception("Duplication échouée")
            return Response({"success": False, "message": str(e)}, status=400)



    # ---------- STATS GLOBAL + PAR CENTRE ----------
    def _detect_date_field(self):
        """
        Détermine le champ date à utiliser pour les stats.
        Priorité : date_debut > start_date > created_at
        """
        model = Formation
        for name in ("date_debut", "start_date", "created_at"):
            try:
                model._meta.get_field(name)
                return name
            except Exception:
                continue
        # Par défaut (sécurité) :
        return "created_at"

    def _stats_from_queryset(self, qs, annee=None):
        """
        Retourne une liste [{ 'mois': 'YYYY-MM', 'total': N }, ...] calculée depuis le queryset.
        """
        date_field = self._detect_date_field()
        if annee:
            qs = qs.filter(**{f"{date_field}__year": int(annee)})
        agg = (
            qs.annotate(mois=TruncMonth(date_field))
              .values("mois")
              .annotate(total=Count("id"))
              .order_by("mois")
        )
        # Sérialisation YYYY-MM
        out = []
        for row in agg:
            m = row["mois"]
            out.append({"mois": f"{m.year:04d}-{m.month:02d}", "total": row["total"]})
        return out

    @extend_schema(summary="Statistiques mensuelles des formations (global + par centre)")
    @action(detail=False, methods=["get"])
    def stats_par_mois(self, request):
        """
        Renvoie:
        - data: stats globales (comportement historique, via Formation.get_stats_par_mois)
        - extra.global_scoped: stats sur le périmètre visible de l'utilisateur
        - extra.par_centre: stats par centre visible (id, nom, stats)
        """
        annee = request.query_params.get("annee")

        # 1) ⚙️ Comportement historique (global 'absolu') – on ne casse rien
        try:
            stats_global = Formation.get_stats_par_mois(annee=annee)
        except Exception as e:
            logger.warning(f"Fallback stats_global via ORM (get_stats_par_mois indisponible): {e}")
            stats_global = self._stats_from_queryset(Formation.objects.all(), annee=annee)

        # 2) 🔒 Périmètre visible (scopé par centres)
        qs_scoped = self.get_queryset()
        stats_global_scoped = self._stats_from_queryset(qs_scoped, annee=annee)

        # 3) 🏢 Stats par centre visible
        centre_rows = (
            qs_scoped.values_list("centre_id", "centre__nom")
            .distinct()
            .order_by("centre__nom")
        )
        par_centre = []
        for cid, cnom in centre_rows:
            if cid is None:
                continue
            stats_c = self._stats_from_queryset(qs_scoped.filter(centre_id=cid), annee=annee)
            par_centre.append({
                "centre_id": cid,
                "centre_nom": cnom,
                "stats": stats_c,
            })

        return Response({
            "success": True,
            "data": stats_global,  # 🔵 inchangé pour compat
            "extra": {
                "global_scoped": stats_global_scoped,  # 🔒 scope user
                "par_centre": par_centre,              # 🏢 détail par centre
            }
        })

    @extend_schema(
        summary="Liste simplifiée des formations (sans pagination)",
        description="Retourne une liste allégée (id, nom, num_offre) de toutes les formations actives, sans pagination."
    )
    @action(detail=False, methods=["get"], url_path="liste-simple")
    def liste_simple(self, request):
        formations = self.get_queryset().only("id", "nom", "num_offre").order_by("nom")  # ✅ restreint
        data = [
            {"id": f.id, "nom": f.nom, "num_offre": getattr(f, "num_offre", None)}
            for f in formations
        ]
        return Response({"success": True, "data": data})

    # ------------------------------------------------------------------
    # 🔹 Archiver / Restaurer (appelle directement les méthodes du modèle)
    # ------------------------------------------------------------------
    @extend_schema(summary="Archiver une formation")
    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        formation = get_object_or_404(Formation.objects.all_including_archived(), pk=pk)
        if formation.est_archivee:
            return Response({"detail": "Déjà archivée."}, status=status.HTTP_400_BAD_REQUEST)

        formation.archiver(user=request.user, commentaire="Archivage manuel via API")
        return Response({"status": "archived"}, status=status.HTTP_200_OK)


    @extend_schema(summary="Restaurer une formation archivée")
    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        formation = get_object_or_404(Formation.objects.all_including_archived(), pk=pk)
        if not formation.est_archivee:
            return Response({"detail": "Déjà active."}, status=status.HTTP_400_BAD_REQUEST)

        formation.desarchiver(user=request.user, commentaire="Restauration manuelle via API")
        return Response({"status": "unarchived"}, status=status.HTTP_200_OK)

        
    @extend_schema(summary="Lister uniquement les formations archivées")
    @action(detail=False, methods=["get"], url_path="archivees")
    def archivees(self, request):
        qs = self._restrict_to_user_centres(Formation.objects.filter(activite="archivee"))
        serializer = self.get_serializer(qs, many=True)
        return Response({
            "success": True,
            "message": "Liste des formations archivées",
            "data": serializer.data
        })


    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte les formations au format Excel (.xlsx)
        Par défaut : exclut les formations archivées.
        Si ?avec_archivees=true (GET) ou {"avec_archivees": true} (POST) → les inclut.
        """
        # ==========================================================
        # 🆕 Inclusion optionnelle des formations archivées
        # ==========================================================
        inclure_archivees = False
        if request.method == "GET":
            inclure_archivees = request.query_params.get("avec_archivees", "false").lower() == "true"
        elif request.method == "POST":
            inclure_archivees = bool(request.data.get("avec_archivees"))

        if inclure_archivees:
            qs = Formation.objects.all_including_archived().select_related("centre", "type_offre", "statut")
            logger.info(f"[EXPORT XLSX] {request.user} a demandé l’export avec formations archivées.")
        else:
            qs = self.get_queryset().select_related("centre", "type_offre", "statut")

        # ==========================================================
        # 🧩 Gestion des IDs envoyés (POST)
        # ==========================================================
        if request.method == "POST":
            ids = request.data.get("ids", [])
            if isinstance(ids, str):
                ids = [int(x) for x in ids.split(",") if x.isdigit()]
            elif isinstance(ids, list):
                ids = [int(x) for x in ids if str(x).isdigit()]
            else:
                ids = []
            if ids:
                qs = qs.filter(id__in=ids)

        # ==========================================================
        # 📅 Date/heure FR (Europe/Paris)
        # ==========================================================
        tz_paris = pytz.timezone("Europe/Paris")
        now_fr = dj_timezone.now().astimezone(tz_paris)

        wb = Workbook()
        ws = wb.active
        ws.title = "Formations"

        # ==========================================================
        # 🖼️ Logo optionnel
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre et date
        # ==========================================================
        ws.merge_cells("B1:Z1")
        ws["B1"] = "Export des formations — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:Z2")
        ws["B2"] = f"Export réalisé le {now_fr.strftime('%d/%m/%Y à %H:%M (%Z)')}"
        ws["B2"].font = Font(italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])

        # ==========================================================
        # 🧩 Indication visible si on inclut les archivées
        # ==========================================================
        if inclure_archivees:
            ws.append(["⚠️ Export incluant les formations archivées"])
            ws["A4"].font = Font(italic=True, color="FF0000")
            ws.append([])

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "ID", "Centre", "Formation", "Activité", "Type d’offre", "Statut", "Statut temporel",
            "Numéro d’offre", "Date début", "Date fin",
            "Assistante",
            "Places CRIF", "Places MP", "Places prévues (total)", "Capacité max",
            "Inscrits CRIF", "Inscrits MP", "Inscrits (total)",
            "Places dispo", "Places restantes CRIF", "Places restantes MP",
            "Taux saturation (%)", "Taux transformation (%)",
            "Nombre de candidats", "Nombre d’entretiens", "Entrées en formation",
            "Dernier commentaire", "Numéro produit", "Numéro Kairos", "Convocation envoyée",
            "Intitulé du diplôme / titre visé", "Code diplôme", "Code RNCP",
            "Durée totale (heures)", "Heures à distance",
            "Est archivée ?",  # ✅ une seule fois, après activité
        ]
        ws.append(headers)
        header_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))

        header_fill = PatternFill("solid", fgColor="B7DEE8")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        # ==========================================================
        # 🧮 Données et styles
        # ==========================================================
        def _fmt(val):
            if val is None:
                return ""
            if isinstance(val, (datetime.date, datetime.datetime)):
                return val.strftime("%d/%m/%Y")
            return val

        even_fill = PatternFill("solid", fgColor="EEF3FF")
        odd_fill = PatternFill("solid", fgColor="FAFBFD")
        numeric_cols = list(range(10, 21)) + [23, 24, 25, 32, 33]

        for i, f in enumerate(qs, start=1):
            # 🔹 Dernier commentaire
            dernier_commentaire = ""
            if hasattr(f, "get_commentaires"):
                try:
                    last_comment = f.get_commentaires(limit=1).first()
                    if last_comment:
                        contenu_html = getattr(last_comment, "contenu", "") or getattr(last_comment, "body", "")
                        contenu_txt = strip_html_tags_pretty(contenu_html)
                        auteur = getattr(last_comment.created_by, "username", "")
                        date = last_comment.created_at.strftime("%d/%m/%Y %H:%M") if last_comment.created_at else ""
                        texte_final = contenu_txt[:200].strip()
                        if len(contenu_txt) > 200:
                            texte_final += "…"
                        dernier_commentaire = f"[{date}] {auteur} : {texte_final}"
                except Exception:
                    dernier_commentaire = ""

            # 🔹 Données formation
            raw_taux = getattr(f, "taux_saturation", 0) or 0
            taux_pct = (raw_taux * 100) if raw_taux <= 1 else float(raw_taux)
            taux_transfo = getattr(f, "taux_transformation", 0) or 0

            est_archivee = getattr(f, "est_archivee", False)
            activite = getattr(f, "activite", "active")
            activite_display = "Archivée" if activite.lower() == "archivee" else "Active"
            is_archived = activite.lower() == "archivee"

            row = [
                f.id,
                getattr(f.centre, "nom", ""),
                f.nom,
                activite_display,
                getattr(f.type_offre, "nom", ""),
                getattr(f.statut, "nom", ""),
                getattr(f, "status_temporel", ""),
                f.num_offre or "",
                _fmt(f.start_date),
                _fmt(f.end_date),
                f.assistante or "",
                f.prevus_crif or 0,
                f.prevus_mp or 0,
                (f.prevus_crif or 0) + (f.prevus_mp or 0),
                f.cap or "",
                f.inscrits_crif or 0,
                f.inscrits_mp or 0,
                (f.inscrits_crif or 0) + (f.inscrits_mp or 0),
                getattr(f, "places_disponibles", 0) or 0,
                getattr(f, "places_restantes_crif", 0) or 0,
                getattr(f, "places_restantes_mp", 0) or 0,
                taux_pct,
                taux_transfo,
                f.nombre_candidats or 0,
                f.nombre_entretiens or 0,
                getattr(f, "entree_formation", 0) or 0,
                dernier_commentaire,
                f.num_produit or "",
                f.num_kairos or "",
                "Oui" if f.convocation_envoie else "Non",
                f.intitule_diplome or "",
                f.code_diplome or "",
                f.code_rncp or "",
                f.total_heures or 0,
                f.heures_distanciel or 0,
                "Oui" if est_archivee else "Non",
            ]
            ws.append(row)

            # 🟫 Couleur de fond selon activité
            if is_archived:
                fill = PatternFill("solid", fgColor="DDDDDD")  # gris clair pour archivées
            else:
                fill = even_fill if i % 2 == 0 else odd_fill

            # ✅ Styles appliqués à chaque cellule
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrapText=True)
                if j in numeric_cols:
                    cell.number_format = "#,##0"
                    cell.font = Font(color="003366")
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[ws.max_row].height = 30

        # ==========================================================
        # 📊 Filtres et largeur des colonnes
        # ==========================================================
        end_row = ws.max_row
        if end_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[letter].width = min(max_len + 3, 42)

        ws.column_dimensions[get_column_letter(len(headers))].width = 80
        ws.oddFooter.center.text = f"© Rap_App — export du {now_fr.strftime('%d/%m/%Y %H:%M (%Z)')}"

        # ==========================================================
        # 📤 Sauvegarde et réponse
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        binary = buffer.getvalue()

        filename = f'formations_{now_fr.strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        response["Content-Length"] = len(binary)
        return response
