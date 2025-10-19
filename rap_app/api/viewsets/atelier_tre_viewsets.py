from django.db.models import Q, Count
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML, CSS

import csv
import logging
from io import BytesIO
from pathlib import Path
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone as dj_timezone

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from ...models.atelier_tre import (
    AtelierTRE,
    PresenceStatut,
)
from ...models.candidat import Candidat
from ..serializers.atelier_tre_serializers import (
    AtelierTRESerializer,
    AtelierTREMetaSerializer,
)
from ..permissions import IsStaffOrAbove, is_staff_or_staffread
from ..paginations import RapAppPagination

logger = logging.getLogger(__name__)


class AtelierTREViewSet(viewsets.ModelViewSet):
    """
    CRUD minimal des ateliers TRE (M2M direct pour les candidats).
    Accès réservé au staff (lecture/écriture).

    - Filtres: type_atelier, centre, date_atelier (exact/gte/lte)
    - Tri: date_atelier, type_atelier, id
    - Endpoints utilitaires: /meta, /add-candidats, /remove-candidats
    - Présences: /set-presences, /mark-present, /mark-absent

    ⚠️ Scope centres :
      - Admin/Superadmin : accès global
      - Staff : limité aux ateliers dont centre ∈ user.centres
    """
    permission_classes = [IsStaffOrAbove]  # ✅ staff-only partout
    pagination_class = RapAppPagination
    serializer_class = AtelierTRESerializer

    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = {
        "type_atelier": ["exact", "in"],
        "centre": ["exact", "isnull"],
        "date_atelier": ["exact", "date", "gte", "lte"],
    }
    ordering_fields = ["date_atelier", "type_atelier", "id"]
    ordering = ["-date_atelier", "-id"]

    # --------------------- helpers scope/permissions ---------------------

    def _is_admin_like(self, user) -> bool:
        """True si admin ou superadmin."""
        return getattr(user, "is_superuser", False) or (
            hasattr(user, "is_admin") and user.is_admin()
        )

    def _staff_centre_ids(self, user):
        """Liste des centres visibles par staff/staff_read (None si admin-like = accès global)."""
        if self._is_admin_like(user):
            return None
        if is_staff_or_staffread(user):  # ✅ inclut staff_read
            return list(user.centres.values_list("id", flat=True))
        return []

    def _scope_qs_to_user_centres(self, qs):
        """Filtre le queryset selon les centres accessibles."""
        user = self.request.user
        if not user.is_authenticated:
            return qs.none()

        # Candidats/stagiaires → pas d'accès
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            return qs.none()

        centre_ids = self._staff_centre_ids(user)

        # admin/superadmin → pas de restriction
        if centre_ids is None:
            return qs

        # staff/staff_read avec centres
        if centre_ids:
            return qs.filter(centre_id__in=centre_ids).distinct()

        # staff/staff_read sans centre → aucun résultat
        return qs.none()

    def _assert_staff_can_use_centre(self, centre):
        """Empêche un staff/staff_read d'écrire hors de son périmètre de centres."""
        if not centre:
            return
        user = self.request.user
        if self._is_admin_like(user):
            return
        if is_staff_or_staffread(user):  # ✅ inclut staff_read
            allowed = set(user.centres.values_list("id", flat=True))
            if getattr(centre, "id", None) not in allowed:
                raise PermissionDenied("Centre hors de votre périmètre.")

    # ------------------------------ queryset ------------------------------

    def get_queryset(self):
        base = (
            AtelierTRE.objects
            .annotate(
                nb_inscrits_calc=Count("candidats", distinct=True),

                # ✅ Compteurs de présences par statut
                pres_present=Count("presences", filter=Q(presences__statut=PresenceStatut.PRESENT), distinct=True),
                pres_absent=Count("presences", filter=Q(presences__statut=PresenceStatut.ABSENT), distinct=True),
                pres_excuse=Count("presences", filter=Q(presences__statut=PresenceStatut.EXCUSE), distinct=True),
                pres_inconnu=Count("presences", filter=Q(presences__statut=PresenceStatut.INCONNU), distinct=True),
            )
            .select_related("centre", "created_by", "updated_by")
            .prefetch_related("candidats", "presences__candidat")
        )
        return self._scope_qs_to_user_centres(base)


    # --- création / mise à jour (propager l'utilisateur si supporté par BaseModel.save) ---

    def perform_create(self, serializer):
        instance = serializer.save()
        # Vérifie le centre (si fourni par le payload)
        self._assert_staff_can_use_centre(getattr(instance, "centre", None))
        try:
            instance.save(user=self.request.user)
        except TypeError:
            instance.save()

    def perform_update(self, serializer):
        # On récupère le centre proposé (ou existant)
        current = serializer.instance
        new_centre = serializer.validated_data.get("centre", getattr(current, "centre", None))
        self._assert_staff_can_use_centre(new_centre)

        instance = serializer.save()
        try:
            instance.save(user=self.request.user)
        except TypeError:
            instance.save()

    # --- Meta (petit durcissement) ---
    @extend_schema(responses=AtelierTREMetaSerializer)
    @action(detail=False, methods=["get"], url_path="meta", url_name="meta", permission_classes=[IsStaffOrAbove])
    def meta(self, request):
        # instancier avec un "instance={}" pour forcer la représentation complète
        ser = AtelierTREMetaSerializer(instance={}, context={"request": request})
        return Response(ser.data)

    # --- Actions candidats (ajout/retrait sans remplacer toute la liste) ------

    @extend_schema(
        request={"application/json": {"type": "object", "properties": {
            "candidats": {"type": "array", "items": {"type": "integer"}}
        }}},
        responses=AtelierTRESerializer,
        summary="Ajouter des candidats",
        description=(
            "Ajoute des candidats (IDs) à l'atelier sans écraser les existants. "
            "Les candidats doivent appartenir au même centre que l’atelier (via candidat.formation.centre)."
        ),
    )
    @action(detail=True, methods=["post"], url_path="add-candidats", permission_classes=[IsStaffOrAbove])
    def add_candidats(self, request, pk=None):
        atelier = self.get_object()  # ✅ déjà scopé par get_queryset()
        ids = request.data.get("candidats", [])
        if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
            return Response(
                {"detail": "'candidats' doit être une liste d'entiers."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not ids:
            return Response(self.get_serializer(atelier).data)

        qs = Candidat.objects.filter(id__in=ids)

        if not qs.exists():
            return Response(
                {"detail": "Aucun candidat trouvé pour les IDs fournis."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ✅ Vérifier appartenance au même centre (via la formation du candidat)
        atelier_centre_id = getattr(atelier.centre, "id", None)
        mismatched = [
            c.id for c in qs
            if getattr(getattr(c, "formation", None), "centre_id", None) != atelier_centre_id
        ]
        if mismatched:
            return Response(
                {"detail": f"Candidats hors centre de l'atelier: {sorted(mismatched)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        atelier.candidats.add(*qs)
        return Response(self.get_serializer(atelier).data)

    @extend_schema(
        request={"application/json": {"type": "object", "properties": {
            "candidats": {"type": "array", "items": {"type": "integer"}}
        }}},
        responses=AtelierTRESerializer,
        summary="Retirer des candidats",
        description="Retire des candidats (IDs) de l'atelier.",
    )
    @action(detail=True, methods=["post"], url_path="remove-candidats", permission_classes=[IsStaffOrAbove])
    def remove_candidats(self, request, pk=None):
        atelier = self.get_object()
        ids = request.data.get("candidats", [])
        if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
            return Response(
                {"detail": "'candidats' doit être une liste d'entiers."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not ids:
            return Response(self.get_serializer(atelier).data)

        qs = Candidat.objects.filter(id__in=ids)
        if not qs.exists():
            return Response(
                {"detail": "Aucun candidat trouvé pour les IDs fournis."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        atelier.candidats.remove(*qs)
        return Response(self.get_serializer(atelier).data)

    # --- Présences ------------------------------------------------------------

    @extend_schema(
        request={
            "application/json": {
                "type": "object",
                "properties": {
                    "items": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "candidat": {"type": "integer"},
                                "statut": {"type": "string"},
                                "commentaire": {"type": "string"},
                            },
                            "required": ["candidat", "statut"],
                        },
                    }
                },
                "required": ["items"],
            }
        },
        responses=AtelierTRESerializer,
        summary="Définir des présences (upsert par candidat)",
        description="Met à jour (ou crée) la présence pour chaque (candidat, atelier).",
    )
    @action(detail=True, methods=["post"], url_path="set-presences", permission_classes=[IsStaffOrAbove])
    def set_presences(self, request, pk=None):
        atelier = self.get_object()
        items = request.data.get("items", [])
        if not isinstance(items, list):
            return Response(
                {"detail": "'items' doit être une liste d'objets {candidat, statut, commentaire?}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowed = {code for code, _ in PresenceStatut.choices}
        pairs = {}
        for it in items:
            if not isinstance(it, dict):
                return Response({"detail": "Chaque item doit être un objet."}, status=status.HTTP_400_BAD_REQUEST)
            cid = it.get("candidat")
            st = it.get("statut")
            com = it.get("commentaire", None)
            if not isinstance(cid, int) or st not in allowed:
                return Response({"detail": f"Item invalide: {it!r}"}, status=status.HTTP_400_BAD_REQUEST)
            pairs[cid] = {"statut": st, "commentaire": com}

        if not pairs:
            return Response(self.get_serializer(atelier).data)

        # Vérifie que les candidats existent
        wanted_ids = set(pairs.keys())
        existing_ids = set(Candidat.objects.filter(id__in=wanted_ids).values_list("id", flat=True))
        unknown = wanted_ids - existing_ids
        if unknown:
            return Response({"detail": f"Candidats introuvables: {sorted(unknown)}"}, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Vérifie appartenance au même centre que l'atelier
        atelier_centre_id = getattr(atelier.centre, "id", None)
        mismatch = [
            cid for cid in wanted_ids
            if getattr(getattr(Candidat.objects.get(id=cid), "formation", None), "centre_id", None) != atelier_centre_id
        ]
        if mismatch:
            return Response(
                {"detail": f"Candidats hors centre de l'atelier: {sorted(mismatch)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Vérifie inscription préalable à l'atelier
        inscrits_ids = set(atelier.candidats.values_list("id", flat=True))
        not_enrolled = wanted_ids - inscrits_ids
        if not_enrolled:
            return Response(
                {"detail": f"Candidats non inscrits à l'atelier: {sorted(not_enrolled)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Upsert
        for cid in wanted_ids:
            data = pairs[cid]
            c = Candidat.objects.get(id=cid)
            atelier.set_presence(candidat=c, statut=data["statut"], commentaire=data.get("commentaire"), user=request.user)

        return Response(self.get_serializer(atelier).data)

    @extend_schema(
        request={"application/json": {"type": "object", "properties": {"candidats": {"type": "array", "items": {"type": "integer"}}}}},
        responses=AtelierTRESerializer,
        summary="Marquer présents",
    )
    @action(detail=True, methods=["post"], url_path="mark-present", permission_classes=[IsStaffOrAbove])
    def mark_present(self, request, pk=None):
        atelier = self.get_object()
        ids = request.data.get("candidats", [])
        if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
            return Response({"detail": "'candidats' doit être une liste d'entiers."},
                            status=status.HTTP_400_BAD_REQUEST)

        # ✅ ne garde que les candidats déjà inscrits
        qs = atelier.candidats.filter(id__in=ids)
        for c in qs:
            # (optionnel) vérifier centre cohérent ici aussi — l'inscription l'a déjà garanti
            atelier.set_presence(c, PresenceStatut.PRESENT, user=request.user)
        return Response(self.get_serializer(atelier).data)

    @extend_schema(
        request={"application/json": {"type": "object", "properties": {"candidats": {"type": "array", "items": {"type": "integer"}}}}},
        responses=AtelierTRESerializer,
        summary="Marquer absents",
    )
    @action(detail=True, methods=["post"], url_path="mark-absent", permission_classes=[IsStaffOrAbove])
    def mark_absent(self, request, pk=None):
        atelier = self.get_object()
        ids = request.data.get("candidats", [])
        if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
            return Response({"detail": "'candidats' doit être une liste d'entiers."},
                            status=status.HTTP_400_BAD_REQUEST)

        # ✅ idem ici
        qs = atelier.candidats.filter(id__in=ids)
        for c in qs:
            atelier.set_presence(c, PresenceStatut.ABSENT, user=request.user)
        return Response(self.get_serializer(atelier).data)
    
    # ------------------------------------
    #      Export des Ateliers TRE
    # ------------------------------------

    @extend_schema(
        summary="Exporter les ateliers TRE (Excel)",
        description=(
            "Exporte la liste filtrée et autorisée des ateliers TRE au format Excel (.xlsx). "
            "Les filtres, tris et permissions sont appliqués comme pour la vue principale."
        ),
        responses={200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}}},
    )
    @action(detail=False, methods=["get"], url_path="export-xlsx", permission_classes=[IsStaffOrAbove])
    def export_xlsx(self, request):
        """
        🔹 Export Excel des ateliers TRE (avec logo, styles, filtres, formats, etc.)
        Inclut les inscrits et les présents (noms concaténés).
        """
        qs = self.filter_queryset(self.get_queryset())
        logger.debug("📤 export XLSX ateliers TRE — params=%s rows=%d", dict(request.query_params), qs.count())

        wb = Workbook()
        ws = wb.active
        ws.title = "Ateliers TRE"

        # ==========================================================
        # 🖼️ Logo Rap_App
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre principal
        # ==========================================================
        ws.merge_cells("B1:N1")
        ws["B1"] = "Export des ateliers TRE — Rap_App"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:N2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append([])

        # Ligne décorative
        sep_row = ws.max_row + 1
        ws.append(["" for _ in range(10)])
        for cell in ws[sep_row]:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
        ws.row_dimensions[sep_row].height = 5
        ws.append([])

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "ID", "Type d’atelier", "Centre", "Date de l’atelier",
            "Nb inscrits", "Présents", "Absents", "Excusés", "Inconnus",
            "Noms inscrits", "Noms présents", "Créé par", "Créé le", "Modifié le",
        ]
        ws.append(headers)

        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        for i, atelier in enumerate(qs, start=1):
            # 🔹 Noms inscrits
            inscrits_qs = getattr(atelier, "candidats", None)
            inscrits = [getattr(c, "nom", "") for c in inscrits_qs.all()] if inscrits_qs else []
            inscrits_txt = ", ".join(sorted(set(inscrits))) or ""

            # 🔹 Noms présents
            pres_qs = getattr(atelier, "presences", None)
            presents = [
                getattr(p.candidat, "nom", "")
                for p in pres_qs.all()
                if getattr(p, "statut", None) == PresenceStatut.PRESENT
            ] if pres_qs else []
            presents_txt = ", ".join(sorted(set(presents))) or ""

            ws.append([
                atelier.id,
                getattr(atelier.type_atelier, "label", str(atelier.type_atelier)),
                getattr(atelier.centre, "nom", ""),
                atelier.date_atelier.strftime("%d/%m/%Y") if atelier.date_atelier else "",
                getattr(atelier, "nb_inscrits_calc", 0),
                getattr(atelier, "pres_present", 0),
                getattr(atelier, "pres_absent", 0),
                getattr(atelier, "pres_excuse", 0),
                getattr(atelier, "pres_inconnu", 0),
                inscrits_txt,
                presents_txt,
                getattr(getattr(atelier, "created_by", None), "username", ""),
                atelier.created_at.strftime("%d/%m/%Y %H:%M") if atelier.created_at else "",
                atelier.updated_at.strftime("%d/%m/%Y %H:%M") if atelier.updated_at else "",
            ])

            fill = even_fill if i % 2 == 0 else odd_fill
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10, color="333333")
                cell.alignment = Alignment(vertical="top", wrapText=True)

            ws.row_dimensions[ws.max_row].height = 22

        # ==========================================================
        # 📊 Filtres + gel d’en-tête
        # ==========================================================
        end_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        if end_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        # ==========================================================
        # 📏 Largeurs optimisées
        # ==========================================================
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            if letter in ["J", "K"]:  # Noms inscrits / présents
                ws.column_dimensions[letter].width = 60
            else:
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)

        # ==========================================================
        # 📈 Pied de page
        # ==========================================================
        ws.append([])
        ws.append([""])
        ws.append([f"Nombre total d’ateliers exportés : {qs.count()}"])
        ws[ws.max_row][0].font = Font(name="Calibri", bold=True, color="004C99", size=11)

        ws.oddFooter.center.text = f"© Rap_App — export généré le {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        # ==========================================================
        # 📤 Génération du fichier
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        binary_content = buffer.getvalue()

        filename = f"ateliers_tre_{dj_timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response
