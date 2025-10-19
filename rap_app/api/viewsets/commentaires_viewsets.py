# rap_app/api/views/commentaires_views.py
import datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils import timezone as dj_timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from drf_spectacular.utils import extend_schema, OpenApiResponse
from weasyprint import HTML


from ...api.paginations import RapAppPagination
from ...api.permissions import IsStaffOrAbove
from ...api.serializers.commentaires_serializers import (
    CommentaireSerializer,
    CommentaireMetaSerializer,
)
from ...models.commentaires import Commentaire
from ...models.logs import LogUtilisateur
from ..roles import is_admin_like, is_staff_or_staffread, staff_centre_ids


class CommentaireViewSet(viewsets.ModelViewSet):
    """
    💬 Gestion des commentaires liés aux formations.
    """

    queryset = (
        Commentaire.objects.select_related(
            "formation",
            "formation__centre",
            "formation__statut",
            "formation__type_offre",
            "created_by",
        )
        .order_by("-created_at")
        .all()
    )
    serializer_class = CommentaireSerializer
    pagination_class = RapAppPagination
    permission_classes = [IsStaffOrAbove]

    # ------------------------------------------------------------------
    # 🔒 Vérification d'accès à une formation
    # ------------------------------------------------------------------
    def _assert_staff_can_use_formation(self, formation):
        """
        Vérifie qu'un utilisateur staff peut bien accéder à la formation.
        Admins ont toujours accès, sinon on restreint par centre.
        """
        if not formation:
            return

        user = self.request.user

        # Admins / superusers : accès complet
        if is_admin_like(user):
            return

        # Staff ou StaffRead : accès limité à leurs centres
        if is_staff_or_staffread(user):
            allowed_centre_ids = staff_centre_ids(user)
            if formation.centre_id not in allowed_centre_ids:
                raise PermissionDenied("Vous n'avez pas accès à cette formation.")
            return

        # Autres rôles : accès interdit
        raise PermissionDenied("Accès interdit à cette formation.")

    # ------------------------------------------------------------------
    # ⚙️ Options de filtres dynamiques
    # ------------------------------------------------------------------
    @extend_schema(summary="Récupérer les options de filtres pour les commentaires")
    @action(detail=False, methods=["get"], url_path="filter-options")
    def filter_options(self, request):
        """
        Retourne les options disponibles pour le panneau de filtres :
        - Centres accessibles (selon droits et commentaires existants)
        - Types d’offre
        - Formations réellement commentées
        - États distincts de formation (si champ `etat` existe)
        - Statuts du commentaire (actif / archivé)
        - Auteurs des commentaires visibles
        """
        from ...models.formations import Formation, Centre, TypeOffre
        from django.contrib.auth import get_user_model

        User = get_user_model()
        user = request.user

        # 🔒 Restreint au périmètre de visibilité de l'utilisateur
        qs = self.get_queryset()

        # --- Formations concernées ---
        formation_ids = qs.values_list("formation_id", flat=True).distinct()
        formations = Formation.objects.filter(id__in=formation_ids)

        # --- Centres réellement liés à ces formations ---
        centre_ids = formations.values_list("centre_id", flat=True).distinct()
        if is_admin_like(user):
            centres = Centre.objects.filter(id__in=centre_ids)
        elif is_staff_or_staffread(user):
            allowed_ids = set(centre_ids).intersection(staff_centre_ids(user))
            centres = Centre.objects.filter(id__in=allowed_ids)
        else:
            centres = Centre.objects.none()

        # --- Types d’offre utilisés ---
        type_offre_ids = formations.values_list("type_offre_id", flat=True).distinct()
        type_offres = TypeOffre.objects.filter(id__in=type_offre_ids)

        # --- États distincts (si champ existe) ---
        formation_etats = []
        if hasattr(Formation, "etat"):
            etats = (
                formations.exclude(etat__isnull=True)
                .exclude(etat="")
                .values_list("etat", flat=True)
                .distinct()
            )
            formation_etats = [{"value": e, "label": str(e).capitalize()} for e in etats]

        # --- Auteurs de commentaires ---
        auteur_ids = qs.values_list("created_by_id", flat=True).distinct()
        auteurs = User.objects.filter(id__in=auteur_ids)
        auteurs_data = [{"id": a.id, "nom": a.get_full_name() or a.username} for a in auteurs]

        # --- Formations pour dropdown ---
        formations_data = [{"id": f.id, "nom": f.nom} for f in formations]

        # --- Statuts possibles ---
        commentaire_statuts = [
            {"id": Commentaire.STATUT_ACTIF, "nom": "Actif"},
            {"id": Commentaire.STATUT_ARCHIVE, "nom": "Archivé"},
        ]

        data = {
            "centres": [{"id": c.id, "nom": c.nom} for c in centres],
            "type_offres": [{"id": t.id, "nom": t.nom} for t in type_offres],
            "formations": formations_data,
            "formation_etats": formation_etats,
            "auteurs": auteurs_data,
            "statuts": commentaire_statuts,
        }

        return Response({
            "success": True,
            "message": "Options de filtres récupérées avec succès.",
            "data": data,
        })

    # ------------------------------------------------------------------
    # 🔒 Scope par rôle utilisateur
    # ------------------------------------------------------------------
    def get_queryset(self):
        """
        Filtrage complet des commentaires :
        - par centre, type d'offre, état / nom de formation, auteur
        - par statut du commentaire (actif / archivé / tous)
        - périmètre utilisateur (admin / staff / staffread)
        """
        qs = super().get_queryset()
        user = self.request.user
        params = self.request.query_params

        # ------------------------------------------------------------------
        # 🔁 Détermination du "scope" de statut principal
        # ------------------------------------------------------------------
        statut_scope = (params.get("statut_id") or params.get("statut") or "").lower().strip()
        if self.action in {"retrieve", "update", "archiver", "desarchiver"}:
            statut_scope = "all"

        # ------------------------------------------------------------------
        # 🎯 Filtres additionnels
        # ------------------------------------------------------------------
        centre_id = params.get("centre_id")
        if centre_id and str(centre_id).isdigit():
            qs = qs.filter(formation__centre_id=int(centre_id))

        type_offre_id = params.get("type_offre_id")
        if type_offre_id and str(type_offre_id).isdigit():
            qs = qs.filter(formation__type_offre_id=int(type_offre_id))

        formation_etat = params.get("formation_etat")
        if formation_etat:
            qs = qs.filter(formation__etat__iexact=formation_etat)

        auteur_id = params.get("auteur_id")
        if auteur_id and str(auteur_id).isdigit():
            qs = qs.filter(created_by_id=int(auteur_id))

        formation_nom = params.get("formation_nom")
        if formation_nom:
            qs = qs.filter(formation__nom__icontains=formation_nom)

        # ------------------------------------------------------------------
        # 🔐 Périmètre utilisateur
        # ------------------------------------------------------------------
        if not is_admin_like(user):
            if is_staff_or_staffread(user):
                centre_ids = staff_centre_ids(user)
                if not centre_ids:
                    return qs.none()
                qs = qs.filter(formation__centre_id__in=centre_ids)
            else:
                return qs.none()

        # ------------------------------------------------------------------
        # 💬 Gestion du statut du commentaire
        # ------------------------------------------------------------------
        if not statut_scope or statut_scope in {"actif", Commentaire.STATUT_ACTIF.lower()}:
            qs = qs.filter(statut_commentaire=Commentaire.STATUT_ACTIF)
        elif statut_scope in {"archive", "archivé", Commentaire.STATUT_ARCHIVE.lower()}:
            qs = qs.filter(statut_commentaire=Commentaire.STATUT_ARCHIVE)
        elif statut_scope == "all":
            pass
        elif statut_scope.isdigit():
            qs = qs.filter(formation__statut_id=int(statut_scope))

        return qs

    # ------------------------------------------------------------------
    # ⚙️ Contexte sérialiseur
    # ------------------------------------------------------------------
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["include_full_content"] = True
        return context

    # ------------------------------------------------------------------
    # 📋 Liste / Lecture
    # ------------------------------------------------------------------
    @extend_schema(summary="Lister les commentaires actifs ou filtrés")
    def list(self, request, *args, **kwargs):
        limit = request.query_params.get("limit")
        queryset = self.filter_queryset(self.get_queryset())

        if limit and str(limit).isdigit():
            queryset = queryset[: int(limit)]

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({"success": True, "message": "Commentaires récupérés", "data": serializer.data})

    @extend_schema(summary="Récupérer un commentaire")
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            "success": True,
            "message": "Commentaire récupéré avec succès.",
            "data": serializer.data,
        })

    # ------------------------------------------------------------------
    # ✏️ Création / Mise à jour / Suppression
    # ------------------------------------------------------------------
    @extend_schema(summary="Créer un commentaire")
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        formation = serializer.validated_data.get("formation")
        self._assert_staff_can_use_formation(formation)

        if formation and hasattr(formation, "saturation"):
            serializer.validated_data["saturation_formation"] = formation.saturation

        commentaire = serializer.save(created_by=request.user)

        LogUtilisateur.log_action(
            instance=commentaire,
            action=LogUtilisateur.ACTION_CREATE,
            user=request.user,
            details=f"Création d'un commentaire pour la formation #{commentaire.formation_id}",
        )

        serializer = self.get_serializer(commentaire)
        return Response(
            {"success": True, "message": "Commentaire créé avec succès.", "data": serializer.data},
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(summary="Mettre à jour un commentaire")
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        new_formation = serializer.validated_data.get("formation", instance.formation)
        self._assert_staff_can_use_formation(new_formation)

        commentaire = serializer.save()

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details=f"Mise à jour du commentaire #{instance.pk}",
        )

        serializer = self.get_serializer(commentaire)
        return Response(
            {"success": True, "message": "Commentaire mis à jour avec succès.", "data": serializer.data},
            status=status.HTTP_200_OK,
        )

    @extend_schema(summary="Supprimer un commentaire")
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self._assert_staff_can_use_formation(getattr(instance, "formation", None))
        instance.delete()

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_DELETE,
            user=request.user,
            details=f"Suppression du commentaire #{instance.pk}",
        )

        return Response(
            {"success": True, "message": "Commentaire supprimé avec succès.", "data": None},
            status=status.HTTP_204_NO_CONTENT,
        )

    # ------------------------------------------------------------------
    # 📊 Statistiques de saturation
    # ------------------------------------------------------------------
    @extend_schema(
        summary="Récupérer les statistiques de saturation des commentaires",
        responses={200: OpenApiResponse(description="Données de saturation pour une formation")},
    )
    @action(detail=False, methods=["get"], url_path="saturation-stats")
    def saturation_stats(self, request):
        formation_id = request.query_params.get("formation_id")
        stats = Commentaire.get_saturation_stats(formation_id=formation_id)
        return Response({
            "success": True,
            "message": "Statistiques de saturation récupérées avec succès.",
            "data": stats,
        })

    # ------------------------------------------------------------------
    # ⚙️ Métadonnées du module commentaire
    # ------------------------------------------------------------------
    @extend_schema(summary="Récupérer les métadonnées des commentaires")
    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        serializer = CommentaireMetaSerializer()
        return Response({
            "success": True,
            "message": "Métadonnées récupérées avec succès.",
            "data": serializer.data,
        })

    # ------------------------------------------------------------------
    # 🗂️ Archivage / Désarchivage logique
    # ------------------------------------------------------------------
    @extend_schema(summary="Archiver un commentaire")
    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        instance = self.get_object()
        instance.archiver(save=True)
        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details=f"Archivage du commentaire #{instance.pk}",
        )
        serializer = self.get_serializer(instance)
        return Response({
            "success": True,
            "message": f"Commentaire #{instance.pk} archivé avec succès.",
            "data": serializer.data,
        })

    @extend_schema(summary="Désarchiver un commentaire")
    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        instance = self.get_object()
        instance.desarchiver(save=True)
        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details=f"Désarchivage du commentaire #{instance.pk}",
        )
        serializer = self.get_serializer(instance)
        return Response({
            "success": True,
            "message": f"Commentaire #{instance.pk} restauré avec succès.",
            "data": serializer.data,
        })



    # ------------------------------------------------------------------
    # 📤 Export PDF / XLSX
    # ------------------------------------------------------------------
    @action(detail=False, methods=["get", "post"], url_path="export")
    def export(self, request):
        fmt = request.data.get("format") or request.query_params.get("format", "pdf")

        if request.method == "POST":
            ids = request.data.get("ids", [])
            export_all = request.data.get("all", False)
        else:
            ids = request.query_params.get("ids", "")
            export_all = request.query_params.get("all", "false").lower() == "true"

        # ✅ Paramètre facultatif pour inclure les archivés
        include_archived = str(
            request.data.get("include_archived")
            or request.query_params.get("include_archived", "")
        ).lower() in {"1", "true", "yes", "on"}

        # -----------------------------------------------------
        # 🔍 Queryset de base
        # -----------------------------------------------------
        qs = self.get_queryset()

        # ✅ Si l'utilisateur demande explicitement d'inclure les archivés
        if include_archived:
            qs = qs.filter(
                statut_commentaire__in=[
                    Commentaire.STATUT_ACTIF,
                    Commentaire.STATUT_ARCHIVE,
                ]
            )
        else:
            qs = qs.filter(statut_commentaire=Commentaire.STATUT_ACTIF)

        # -----------------------------------------------------
        # 🎯 Filtrage par IDs sélectionnés
        # -----------------------------------------------------
        if ids:
            if isinstance(ids, str):
                id_list = [int(i) for i in ids.split(",") if i.isdigit()]
            else:
                id_list = [int(i) for i in ids if str(i).isdigit()]
            qs = qs.filter(id__in=id_list)

        if not export_all and not ids:
            return Response({"detail": "Aucun commentaire sélectionné"}, status=400)

        # -----------------------------------------------------
        # 📄 Format d’export
        # -----------------------------------------------------
        if fmt == "pdf":
            return self._export_pdf(qs)
        elif fmt == "xlsx":
            return self._export_xlsx(qs)
        return Response({"detail": "Format non supporté (seuls pdf, xlsx)"}, status=400)

    # ------------------------------------------------------------------
    # 📄 Génération PDF — avec logo, titre, activité et pied de page
    # ------------------------------------------------------------------
    def _export_pdf(self, qs):
        """
        Génère un PDF avec logo, titre, activité, date et pied de page.
        Compatible dev et prod.
        """
        data = []
        for c in qs:
            f = getattr(c, "formation", None)
            data.append({
                "id": c.id,
                "contenu": getattr(c, "contenu", "") or "",
                "auteur": getattr(c.created_by, "username", ""),
                "activite": getattr(c, "activite", "") or "",
                "created_at": c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "",
                "saturation_formation": c.saturation_formation or c.saturation,
                "taux_saturation": getattr(f, "taux_saturation", ""),
                "formation": {
                    "nom": getattr(f, "nom", "") if f else "",
                    "num_offre": getattr(f, "num_offre", "") if f else "",
                    "centre_nom": getattr(f.centre, "nom", "") if f and f.centre else "",
                    "type_offre_nom": getattr(f.type_offre, "nom", "") if f and f.type_offre else "",
                    "statut_nom": getattr(f.statut, "nom", "") if f and f.statut else "",
                    "saturation_commentaires": (
                        f.get_saturation_moyenne_commentaires() if f else ""
                    ),
                    # 🧼 Le contenu HTML est déjà nettoyé via bleach côté serializer,
                    # donc on peut l'inclure tel quel dans le template PDF.

                },
            })

        request = self.request
        # 🔹 Logo visible en dev et prod
        try:
            logo_url = request.build_absolute_uri(static("images/logo.png"))
        except Exception:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            logo_url = f"file://{logo_path}"

        context = {
            "items": data,
            "user": request.user,
            "now": dj_timezone.now(),
            "logo_url": logo_url,
        }

        html_string = render_to_string("exports/commentaires_pdf.html", context)
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

        filename = f'commentaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        response["Content-Length"] = len(pdf)
        return response


    # ------------------------------------------------------------------
    # 📊 Génération Excel — avec logo, titre, activité et date d’export
    # ------------------------------------------------------------------
    def _export_xlsx(self, qs):
        """
        Génère un fichier Excel stylisé avec logo, titre, activité,
        date d’export et ajustement automatique des colonnes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Commentaires Formation"

        # ==========================================================
        # 🖼️ Logo Rap_App
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre + date d’export
        # ==========================================================
        ws.merge_cells("B1:M1")
        ws["B1"] = "Commentaires et plan d’action — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="0077CC")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:M2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])  # ligne vide

        # ==========================================================
        # 📋 En-têtes des colonnes
        # ==========================================================
        headers = [
            "ID", "Contenu", "Activité", "Auteur", "Créé le",
            "Formation", "N° Offre", "Centre", "Type d’offre", "Statut",
            "Saturation au moment du commentaire (%)",
            "Saturation actuelle (%)",
            "Moy. des commentaires (%)",
        ]
        ws.append(headers)

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="E9F2FF")

        def _fmt(val):
            if val is None:
                return ""
            if isinstance(val, datetime.datetime):
                return val.strftime("%d/%m/%Y %H:%M")
            return str(val)

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        for c in qs:
            f = getattr(c, "formation", None)
            ws.append([
                c.id,
                c.contenu or "",
                getattr(c, "activite", "") or "",
                getattr(c.created_by, "username", ""),
                _fmt(c.created_at),
                getattr(f, "nom", "") if f else "",
                getattr(f, "num_offre", "") if f else "",
                getattr(f.centre, "nom", "") if f and f.centre else "",
                getattr(f.type_offre, "nom", "") if f and f.type_offre else "",
                getattr(f.statut, "nom", "") if f and f.statut else "",
                c.saturation_formation or c.saturation,
                getattr(f, "taux_saturation", ""),
                f.get_saturation_moyenne_commentaires() if f else "",
            ])

        # ==========================================================
        # 📏 Ajustement largeur colonnes + wrap sur contenu
        # ==========================================================
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in {"B", "C"}:  # Contenu + Activité
                ws.column_dimensions[col_letter].width = 50
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            else:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        # ==========================================================
        # 📤 Sauvegarde et réponse HTTP
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        binary = buffer.getvalue()

        filename = f'commentaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        response["Content-Length"] = len(binary)
        return response
