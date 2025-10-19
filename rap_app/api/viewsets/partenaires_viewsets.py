from rest_framework import viewsets, status, filters
from rest_framework.response import Response
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiResponse
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import BasePermission, SAFE_METHODS
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, filters as dj_filters
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone as dj_timezone
from django.conf import settings
from pathlib import Path
from io import BytesIO
import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage

from ...api.permissions import IsOwnerOrStaffOrAbove, UserVisibilityScopeMixin, is_staff_or_staffread
from ...models.partenaires import Partenaire
from ...models.logs import LogUtilisateur
from ..serializers.partenaires_serializers import PartenaireChoicesResponseSerializer, PartenaireSerializer



# -------------------- permission locale --------------------

class PartenaireAccessPermission(BasePermission):
    """
    - Admin/staff/superuser : OK
    - Sinon :
        - OK si créateur
        - OK en LECTURE si attribué via une prospection (prospections.owner = user)
    Les droits d'édition/suppression restent limités par la vue (créateur uniquement).
    """
    message = "Accès restreint."

    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)

    def has_object_permission(self, request, view, obj):
        user = request.user
        if getattr(user, "is_superuser", False) or is_staff_or_staffread(user) or (
            hasattr(user, "is_admin") and callable(user.is_admin) and user.is_admin()
        ):
            return True

        if getattr(obj, "created_by_id", None) == user.id:
            return True

        if request.method in SAFE_METHODS and hasattr(obj, "prospections"):
            try:
                return obj.prospections.filter(owner=user).exists()
            except Exception:
                return False

        return False


class InlinePartenaireFilter(FilterSet):
    type = dj_filters.CharFilter(lookup_expr="exact")
    is_active = dj_filters.BooleanFilter()
    city = dj_filters.CharFilter(lookup_expr="icontains")
    secteur_activite = dj_filters.CharFilter(lookup_expr="icontains")
    created_by = dj_filters.NumberFilter()
    # ✅ filtrer par centre par défaut du partenaire
    centre_id = dj_filters.NumberFilter(field_name="default_centre_id")

    has_appairages = dj_filters.BooleanFilter(method="filter_has_appairages")
    has_prospections = dj_filters.BooleanFilter(method="filter_has_prospections")
    has_formations = dj_filters.BooleanFilter(method="filter_has_formations")
    has_candidats = dj_filters.BooleanFilter(method="filter_has_candidats")

    class Meta:
        model = Partenaire
        fields = [
            "type",
            "is_active",
            "city",
            "secteur_activite",
            "created_by",
            "centre_id",
            "has_appairages",
            "has_prospections",
            "has_formations",
            "has_candidats",
        ]

    def _bool(self, value):
        if isinstance(value, bool):
            return value
        if value is None:
            return None
        s = str(value).strip().lower()
        return s in {"1", "true", "yes", "y", "on"}

    def filter_has_appairages(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(appairages_count__gt=0) if b else queryset.filter(appairages_count=0)

    def filter_has_prospections(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(prospections_count__gt=0) if b else queryset.filter(prospections_count=0)

    def filter_has_formations(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(formations_count__gt=0) if b else queryset.filter(formations_count=0)

    def filter_has_candidats(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(candidats_count__gt=0) if b else queryset.filter(candidats_count=0)


@extend_schema_view(
    list=extend_schema(
        summary="Lister les partenaires",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(response=PartenaireSerializer)}
    ),
    retrieve=extend_schema(
        summary="Détail d’un partenaire",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(response=PartenaireSerializer)}
    ),
    create=extend_schema(
        summary="Créer un partenaire",
        tags=["Partenaires"],
        responses={201: OpenApiResponse(description="Création réussie")}
    ),
    update=extend_schema(
        summary="Modifier un partenaire",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(description="Mise à jour réussie")}
    ),
    destroy=extend_schema(
        summary="Supprimer un partenaire",
        tags=["Partenaires"],
        responses={204: OpenApiResponse(description="Suppression réussie")}
    ),
)
class PartenaireViewSet(UserVisibilityScopeMixin, viewsets.ModelViewSet):
    serializer_class = PartenaireSerializer
    # ✅ utilise la permission locale pour autoriser la lecture des partenaires attribués via prospection
    permission_classes = [PartenaireAccessPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = InlinePartenaireFilter
    ordering_fields = ["nom", "created_at", "default_centre__nom"]  # ✅ tri centre
    ordering = ["nom"]
    search_fields = [
        "nom",
        "secteur_activite",
        "street_name",
        "zip_code",
        "city",
        "country",
        "contact_nom",
        "contact_poste",
        "contact_email",
        "contact_telephone",
        "website",
        "social_network_url",
        "description",
        "action_description",
        "actions",
        "created_by__first_name",
        "created_by__last_name",
        "created_by__username",
        "default_centre__nom",  # ✅ recherche centre
    ]

    # -------------------- helpers scope --------------------

    def _is_admin_like(self, user) -> bool:
        return getattr(user, "is_superuser", False) or (hasattr(user, "is_admin") and user.is_admin())

    def _staff_centre_ids(self, user):
        if self._is_admin_like(user):
            return None  # accès global
        if is_staff_or_staffread(user):
            return list(user.centres.values_list("id", flat=True))
        return []  # non-staff

    def _scoped_for_user(self, qs, user):
        """
        Staff :
        - partenaires liés à AU MOINS une formation d'un de ses centres (via appairages/prospections)
        - OU partenaires dont le default_centre ∈ ses centres
        - OU partenaires qu'il a créés
        ⚠️ Exclut les partenaires sans centre si non créés par lui.
        """
        centre_ids = self._staff_centre_ids(user)

        # Cas admin/superadmin : accès global
        if centre_ids is None:
            return qs

        # Cas staff sans centre rattaché → uniquement ses créations
        if not centre_ids:
            return qs.filter(created_by=user)

        # Scoping principal
        scoped = qs.filter(
            Q(appairages__formation__centre_id__in=centre_ids)
            | Q(prospections__formation__centre_id__in=centre_ids)
            | Q(default_centre_id__in=centre_ids)
            | Q(created_by=user)
        ).distinct()

        # 🧩 Debug optionnel (protégé même si DEBUG n’existe pas)
        if getattr(settings, "DEBUG", False):
            orphaned = scoped.filter(default_centre__isnull=True)
            (
                f"[DEBUG] User={user} voit {scoped.count()} partenaires, "
                f"dont {orphaned.count()} sans centre."
            )

        # 🔒 Exclut explicitement les partenaires sans centre,
        # sauf s’ils ont été créés par le staff lui-même
        return scoped.exclude(
            Q(default_centre_id__isnull=True)
            & ~Q(created_by=user)
            & ~Q(appairages__formation__centre_id__in=centre_ids)
            & ~Q(prospections__formation__centre_id__in=centre_ids)
        ).distinct()


    def _user_can_access_partenaire(self, partenaire, user) -> bool:
        """Vérifie qu'un staff est dans le périmètre du partenaire."""
        if self._is_admin_like(user):
            return True
        if not is_staff_or_staffread(user):
            # non-staff : déjà géré par permission/queryset
            return True
        centre_ids = set(user.centres.values_list("id", flat=True))
        if not centre_ids:
            return partenaire.created_by_id == user.id
        linked = (
            partenaire.appairages.filter(formation__centre_id__in=centre_ids).exists()
            or partenaire.prospections.filter(formation__centre_id__in=centre_ids).exists()
            or (partenaire.default_centre_id in centre_ids)
        )
        return linked or (partenaire.created_by_id == user.id)



    # -------------------- queryset --------------------

    def get_queryset(self):
        user = self.request.user

        qs = (
            Partenaire.objects
            .filter(is_active=True)
            .select_related("created_by", "default_centre")   # ✅ pas de N+1 centre
            .annotate(
                prospections_count=Count("prospections", distinct=True),
                appairages_count=Count("appairages", distinct=True),
                # Formations (distinct) via appairages + prospections
                formations_count=Count(
                    "appairages__formation",
                    filter=Q(appairages__formation__isnull=False),
                    distinct=True,
                ) + Count(
                    "prospections__formation",
                    filter=Q(prospections__formation__isnull=False),
                    distinct=True,
                ),
                # Candidats distincts (via appairages)
                candidats_count=Count("appairages__candidat", distinct=True),
            )
        )

        if self._is_admin_like(user):
            return qs
        if is_staff_or_staffread(user):
            return self._scoped_for_user(qs, user)

        # ✅ Candidat·e : créés par lui/elle + attribués via prospection (owner=user)
        return qs.filter(Q(created_by=user) | Q(prospections__owner=user)).distinct()

    # -------------------- endpoints utilitaires --------------------

    @action(detail=False, methods=["get"], url_path="choices")
    def choices(self, request):
        types = [{"value": k, "label": v} for k, v in Partenaire.TYPE_CHOICES]
        actions = [{"value": k, "label": v} for k, v in Partenaire.CHOICES_TYPE_OF_ACTION]
        ser = PartenaireChoicesResponseSerializer(instance={"types": types, "actions": actions})
        return Response(ser.data)

    @extend_schema(summary="🔽 Filtres disponibles pour les partenaires", tags=["Partenaires"])
    @action(detail=False, methods=["get"], url_path="filter-options")
    def filter_options(self, request):
        """
        Options de filtre basées sur le queryset **déjà scopé** de l'utilisateur.
        """
        qs = self.filter_queryset(self.get_queryset())

        villes = (
            qs.exclude(city__isnull=True)
              .exclude(city="")  # ✅ évite l’avertissement Pylance
              .values_list("city", flat=True)
              .distinct()
        )
        secteurs = (
            qs.exclude(secteur_activite__isnull=True)
              .exclude(secteur_activite="")  # ✅ évite l’avertissement Pylance
              .values_list("secteur_activite", flat=True)
              .distinct()
        )
        users = (
            qs.exclude(created_by__isnull=True)
              .values("created_by").distinct()
              .values("created_by", "created_by__first_name", "created_by__last_name")
        )
        # ✅ centres par défaut disponibles
        centres = (
            qs.filter(default_centre__isnull=False)
              .values("default_centre_id", "default_centre__nom")
              .distinct()
        )

        return Response({
            "cities": [{"value": v, "label": v} for v in villes],
            "secteurs": [{"value": s, "label": s} for s in secteurs],
            "users": [
                {
                    "id": u["created_by"],
                    "full_name": " ".join(filter(None, [u.get("created_by__first_name"), u.get("created_by__last_name")])).strip()
                }
                for u in users if u["created_by"]
            ],
            "centres": [
                {"id": c["default_centre_id"], "nom": c["default_centre__nom"]}
                for c in centres
            ],
        })

    # -------------------- CRUD --------------------

    def perform_create(self, serializer):
        user = self.request.user
        default_centre = None

        # 🔎 Cas 1 — Superadmin / Admin : centre optionnel
        if self._is_admin_like(user):
            default_centre = serializer.validated_data.get("default_centre")

        # 🔎 Cas 2 — Staff / Staff_read : doit avoir un centre autorisé
        elif is_staff_or_staffread(user):
            centres = list(user.centres.all())
            if not centres:
                raise PermissionDenied("❌ Vous n’êtes rattaché à aucun centre.")

            # Si un centre est envoyé → vérifie qu’il fait partie des siens
            default_centre = serializer.validated_data.get("default_centre") or centres[0]
            if default_centre and default_centre not in centres:
                raise PermissionDenied("❌ Ce centre n’est pas dans votre périmètre autorisé.")

        # 🔎 Cas 3 — Candidat / Stagiaire : centre = celui de sa formation
        else:
            default_centre = (
                serializer.validated_data.get("default_centre")
                or getattr(user, "centre", None)
            )
            if not default_centre:
                raise PermissionDenied("❌ Impossible de créer un partenaire sans centre associé.")

        # ✅ Vérification finale de sécurité
        if not default_centre:
            raise PermissionDenied("❌ Le centre associé est obligatoire pour créer un partenaire.")

        # ✅ Création de l’objet
        instance = serializer.save(created_by=user, default_centre=default_centre)

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_CREATE,
            user=user,
            details="Création d'un partenaire",
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        user = request.user

        # 🔒 Vérifications d’accès
        if not is_staff_or_staffread(user) and not getattr(user, "is_superuser", False):
            if instance.created_by_id != user.id:
                raise PermissionDenied("Vous ne pouvez modifier que vos propres partenaires.")

        if is_staff_or_staffread(user) and not self._is_admin_like(user):
            if not self._user_can_access_partenaire(instance, user):
                raise PermissionDenied("Partenaire hors de votre périmètre (centres).")

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # ⚙️ Logique ajoutée : remplir le centre manquant automatiquement
        default_centre = instance.default_centre  # valeur actuelle
        if not default_centre:
            if self._is_admin_like(user):
                # Admin : ne rien forcer
                default_centre = serializer.validated_data.get("default_centre")
            elif is_staff_or_staffread(user):
                centres = list(user.centres.all())
                if not centres:
                    raise PermissionDenied("Vous n’êtes rattaché à aucun centre.")
                default_centre = serializer.validated_data.get("default_centre") or centres[0]
            else:
                default_centre = serializer.validated_data.get("default_centre") or getattr(user, "centre", None)
                if not default_centre:
                    raise PermissionDenied("Impossible de modifier un partenaire sans centre associé.")

        instance = serializer.save(default_centre=default_centre)

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details="Modification d'un partenaire",
        )

        return Response(self.get_serializer(instance).data, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()  # respecte le scope
        user = request.user

        if not is_staff_or_staffread(user) and not getattr(user, "is_superuser", False):
            if instance.created_by_id != user.id:
                raise PermissionDenied("Vous ne pouvez supprimer que vos propres partenaires.")

        if is_staff_or_staffread(user) and not self._is_admin_like(user):
            if not self._user_can_access_partenaire(instance, user):
                raise PermissionDenied("Partenaire hors de votre périmètre (centres).")

        instance.is_active = False
        instance.save()
        LogUtilisateur.log_action(
            instance=instance, action=LogUtilisateur.ACTION_DELETE,
            user=request.user, details="Suppression logique d'un partenaire"
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    # -------------------- détail enrichi --------------------

    @extend_schema(
        summary="Détail d’un partenaire avec relations",
        description="Statistiques sur prospections, formations (via appairages/prospections), appairages et candidats.",
        tags=["Partenaires"],
        responses={200: PartenaireSerializer}
    )
    @action(detail=True, methods=["get"], url_path="with-relations")
    def retrieve_with_relations(self, request, pk=None):
        partenaire = self.get_object()  # déjà scopé
        data = self.get_serializer(partenaire).data

        # Harmonisation: renvoie des objets {count}
        data["prospections"] = {"count": partenaire.prospections.count()}
        data["appairages"] = {"count": partenaire.appairages.count()}

        # Formations distinctes via appairages + prospections
        app_ids = set(
            partenaire.appairages.filter(formation__isnull=False).values_list("formation_id", flat=True)
        )
        pros_ids = set(
            partenaire.prospections.filter(formation__isnull=False).values_list("formation_id", flat=True)
        )
        data["formations"] = {"count": len(app_ids.union(pros_ids))}

        # Candidats distincts
        data["candidats"] = {"count": partenaire.appairages.values("candidat_id").distinct().count()}
        return Response(data)

# -------------------- Export Excel --------------------

    @extend_schema(summary="Exporter les partenaires au format XLSX")
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte la liste des partenaires (avec logo, titre, date d’export et mise en forme).
        """
        qs = self.filter_queryset(
            self.get_queryset().select_related("default_centre", "created_by")
        )

        # ==========================================================
        # 📘 Création du classeur
        # ==========================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Partenaires"

        # ==========================================================
        # 🖼️ Logo Rap_App (affiche en dev et prod)
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre + date d’export
        # ==========================================================
        ws.merge_cells("B1:AJ1")
        ws["B1"] = "Export des partenaires — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="0077CC")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:AJ2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])  # ligne vide

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            # Identité
            "ID", "Nom", "Type", "Secteur d’activité",
            # Adresse
            "Numéro de rue", "Adresse", "Complément", "Code postal", "Ville", "Pays",
            # Coordonnées
            "Téléphone général", "Email général",
            # Contact principal
            "Contact nom", "Contact poste", "Contact email", "Contact téléphone",
            # Employeur
            "SIRET", "Type employeur", "Employeur spécifique", "Code APE",
            "Effectif total", "IDCC", "Assurance chômage spéciale",
            # Maître 1
            "Maître 1 - Nom de naissance", "Maître 1 - Prénom", "Maître 1 - Date de naissance",
            "Maître 1 - Courriel", "Maître 1 - Emploi occupé",
            "Maître 1 - Diplôme/titre le plus élevé", "Maître 1 - Niveau diplôme/titre",
            # Maître 2
            "Maître 2 - Nom de naissance", "Maître 2 - Prénom", "Maître 2 - Date de naissance",
            "Maître 2 - Courriel", "Maître 2 - Emploi occupé",
            "Maître 2 - Diplôme/titre le plus élevé", "Maître 2 - Niveau diplôme/titre",
            # Web & Actions
            "Site web", "Réseau social", "Type d’action", "Description action", "Description générale",
            # Métadonnées
            "Slug", "Centre par défaut", "Créé par", "Date création",
            # Statistiques
            "Nb prospections", "Nb formations", "Nb appairages",
        ]
        ws.append(headers)

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="E9F2FF")

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        def _fmt(val):
            if val is None:
                return ""
            if isinstance(val, datetime.datetime):
                return val.strftime("%d/%m/%Y %H:%M")
            if isinstance(val, datetime.date):
                return val.strftime("%d/%m/%Y")
            return str(val)

        for p in qs:
            ws.append([
                # Identité
                p.id, p.nom, p.get_type_display(), p.secteur_activite or "",
                # Adresse
                p.street_number or "", p.street_name or "", p.street_complement or "",
                p.zip_code or "", p.city or "", p.country or "",
                # Coordonnées générales
                p.telephone or "", p.email or "",
                # Contact
                p.contact_nom or "", p.contact_poste or "", p.contact_email or "", p.contact_telephone or "",
                # Employeur
                p.siret or "",
                p.get_type_employeur_display() if p.type_employeur else "",
                p.employeur_specifique or "",
                p.code_ape or "",
                p.effectif_total or "",
                p.idcc or "",
                "Oui" if p.assurance_chomage_speciale else "Non",
                # Maître 1
                p.maitre1_nom_naissance or "",
                p.maitre1_prenom or "",
                _fmt(p.maitre1_date_naissance),
                p.maitre1_courriel or "",
                p.maitre1_emploi_occupe or "",
                p.maitre1_diplome_titre or "",
                p.maitre1_niveau_diplome or "",
                # Maître 2
                p.maitre2_nom_naissance or "",
                p.maitre2_prenom or "",
                _fmt(p.maitre2_date_naissance),
                p.maitre2_courriel or "",
                p.maitre2_emploi_occupe or "",
                p.maitre2_diplome_titre or "",
                p.maitre2_niveau_diplome or "",
                # Web & Actions
                p.website or "",
                p.social_network_url or "",
                p.get_actions_display() if p.actions else "",
                p.action_description or "",
                p.description or "",
                # Métadonnées
                p.slug or "",
                getattr(p.default_centre, "nom", ""),
                getattr(p.created_by, "username", ""),
                _fmt(p.created_at),
                # Stats
                p.nb_prospections, p.nb_formations, p.nb_appairages,
            ])

        # ==========================================================
        # 📏 Largeurs colonnes + wrap sur texte long
        # ==========================================================
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in ["AV", "AW", "AX", "AY"]:  # Descriptions longues
                ws.column_dimensions[col_letter].width = 80
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            else:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        # ==========================================================
        # 📤 Sauvegarde et réponse HTTP
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        binary_content = buffer.getvalue()

        filename = f'partenaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response