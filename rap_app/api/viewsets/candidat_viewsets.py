from rest_framework import viewsets, filters
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from django.conf import settings
from django.utils import timezone as dj_timezone
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema
from rest_framework.renderers import JSONRenderer
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Count, OuterRef, Subquery, IntegerField, Value, Prefetch
from django.template.loader import render_to_string
from weasyprint import HTML, CSS

import csv
import logging
from django.db.models.functions import Coalesce
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from ..roles import is_admin_like, is_staff_or_staffread, staff_centre_ids
from ...models import atelier_tre

# ✅ imports modèles
from ...models.candidat import (
    Candidat,
    HistoriquePlacement,
    ResultatPlacementChoices,
    NIVEAU_CHOICES,
)
from ...models.prospection import Prospection
from ...models.centres import Centre
from ...models.formations import Formation

# ✅ imports serializers
from ..serializers.candidat_serializers import (
    CandidatLiteSerializer,
    CandidatSerializer,
    CandidatListSerializer,
    CandidatCreateUpdateSerializer,
    CandidatQueryParamsSerializer,  # pour valider/normaliser les query params
)

from ..permissions import IsStaffOrAbove
from ..paginations import RapAppPagination
from ...utils.filters import CandidatFilter

# ✅ logger dédié
logger = logging.getLogger("rap_app.candidats")

SENSITIVE_KEYS = {"password", "token", "secret", "api_key", "auth", "credential", "authorization"}


def _sanitize_dict(d: dict) -> dict:
    out = {}
    for k, v in d.items():
        if any(s in k.lower() for s in SENSITIVE_KEYS):
            out[k] = "***"
        else:
            out[k] = v
    return out


# ─────────────────────────────────────────────────────────────────────────────
# ✅ Construction robuste du payload /candidats/meta/ (scope staff inclus)
# ─────────────────────────────────────────────────────────────────────────────
def _build_candidat_meta(user=None) -> dict:
    """Construit les métadonnées pour /candidats/meta en respectant le scope de l’utilisateur."""

    if is_admin_like(user):
        centres_qs = Centre.objects.order_by("nom").only("id", "nom")
        formations_qs = (
            Formation.objects.select_related("centre")
            .only("id", "nom", "num_offre", "centre__nom")
            .order_by("nom")
        )
    elif is_staff_or_staffread(user):
        centre_ids = staff_centre_ids(user) or []
        centres_qs = (
            Centre.objects.filter(id__in=centre_ids)
            .order_by("nom")
            .only("id", "nom")
        )
        formations_qs = (
            Formation.objects.select_related("centre")
            .filter(centre_id__in=centre_ids)
            .only("id", "nom", "num_offre", "centre__nom")
            .order_by("nom")
        )
    else:
        centres_qs = Centre.objects.none()
        formations_qs = Formation.objects.none()

    return {
        "statut_choices": [{"value": k, "label": v} for k, v in Candidat.StatutCandidat.choices],
        "cv_statut_choices": [{"value": k, "label": v} for k, v in Candidat.CVStatut.choices],
        "type_contrat_choices": [{"value": k, "label": v} for k, v in Candidat.TypeContrat.choices],
        "disponibilite_choices": [{"value": k, "label": v} for k, v in Candidat.Disponibilite.choices],
        "resultat_placement_choices": [{"value": k, "label": v} for k, v in ResultatPlacementChoices.choices],
        "contrat_signe_choices": [{"value": k, "label": v} for k, v in Candidat.ContratSigne.choices],
        "niveau_choices": [{"value": val, "label": f"{val} ★"} for val, _ in NIVEAU_CHOICES],
        "centre_choices": [{"value": c.id, "label": c.nom} for c in centres_qs],
        "formation_choices": [
            {
                "value": f.id,
                "label": f"{f.nom}" + (f" — {f.num_offre}" if f.num_offre else ""),
            }
            for f in formations_qs
        ],
    }

class CandidatViewSet(viewsets.ModelViewSet):
    permission_classes = [IsStaffOrAbove]
    pagination_class = RapAppPagination

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = CandidatFilter

    search_fields = [
        "nom",
        "prenom",
        "email",
        "telephone",
        "ville",
        "code_postal",
        "origine_sourcing",
        "numero_osia",
        "formation__nom",
        "formation__num_offre",
        "formation__centre__nom",
        "compte_utilisateur__email",
        "placement_appairage__partenaire__nom",
        "cv_statut",
    ]

    ordering_fields = [
        "date_inscription",
        "nom",
        "prenom",
        "statut",
        "cv_statut",
        "formation",
        "formation__nom",
        "formation__centre__nom",
        "nb_appairages_calc",
        "nb_prospections_calc",
        "date_placement",
        "resultat_placement",
        "contrat_signe",
        "entreprise_placement",
        "entreprise_validee",
        "vu_par",
        "placement_appairage__date_appairage",
        "placement_appairage__partenaire__nom",
    ]
    ordering = ["-date_inscription"]

    # ---------- UTILS LOGGING ----------
    def _qp_dict(self, request):
        qp = {}
        for k in request.query_params.keys():
            vals = request.query_params.getlist(k)
            qp[k] = vals if len(vals) > 1 else (vals[0] if vals else None)
        return _sanitize_dict(qp)

    def _log_filters(self, request, base_qs, filtered_qs):
        try:
            logger.debug("🔎 query_params=%s", self._qp_dict(request))
            before_count = base_qs.count()
            after_count = filtered_qs.count()
            logger.debug("📊 queryset counts: before=%s after=%s", before_count, after_count)
            try:
                logger.debug("🧠 SQL: %s", str(filtered_qs.query))
            except Exception:
                logger.debug("🧠 SQL: <unavailable>")

            backend = DjangoFilterBackend()
            fs = backend.get_filterset(request, base_qs, self)
            if fs is not None:
                valid = fs.is_valid()
                form = getattr(fs, "form", None)
                errors = getattr(form, "errors", {})
                cleaned = getattr(form, "cleaned_data", {})
                logger.debug("🧪 FilterSet valid=%s errors=%s cleaned=%s", valid, errors, cleaned)
        except Exception:
            logger.exception("Erreur pendant le logging des filtres.")

# ---------- helpers scope/permission ----------

    def _scope_qs_to_user_centres(self, qs):
        """
        Staff : ne voit que les candidats dont formation.centre_id ∈ ses centres.
        Admin/superadmin : global.
        """
        user = self.request.user
        if is_admin_like(user):
            return qs  # admin-like → accès global

        centre_ids = staff_centre_ids(user) or []
        if centre_ids:
            return qs.filter(formation__centre_id__in=centre_ids)
        return qs.none()

    def _assert_staff_can_use_formation(self, formation):
        """Empêche un staff d'assigner une formation hors de son périmètre."""
        if not formation:
            return
        user = self.request.user
        if is_admin_like(user):
            return
        if is_staff_or_staffread(user):
            allowed = set(user.centres.values_list("id", flat=True))
            if getattr(formation, "centre_id", None) not in allowed:
                raise PermissionDenied("Formation hors de votre périmètre (centre).")
        # ---------- queryset de base + annotations ----------

 # ---------- queryset de base + annotations ----------
    def base_queryset(self):
        qs = (
            Candidat.objects
            .select_related(
                "formation",
                "formation__centre",
                "formation__type_offre",
                "evenement",
                "compte_utilisateur",
                "responsable_placement",
                "vu_par",
                "entreprise_placement",
                "entreprise_validee",
                "placement_appairage",
                "placement_appairage__partenaire",
                "placement_appairage__created_by",
                "placement_appairage__updated_by",
            )
            .prefetch_related(
                "appairages",
                Prefetch(
                    "ateliers_tre",
                    queryset=atelier_tre.AtelierTRE.objects.only("id", "type_atelier"),
                ),
            )
        )

        # nb d'appairages par candidat (distinct)
        qs = qs.annotate(nb_appairages_calc=Count("appairages", distinct=True))

        # nb de prospections via subquery sur le propriétaire (compte_utilisateur)
        prospection_cnt = (
            Prospection.objects
            .filter(owner_id=OuterRef("compte_utilisateur_id"))
            .values("owner_id")
            .annotate(c=Count("id"))
            .values("c")[:1]
        )
        qs = qs.annotate(
            nb_prospections_calc=Coalesce(
                Subquery(prospection_cnt, output_field=IntegerField()),
                Value(0),
                output_field=IntegerField(),
            )
        )

        # (optionnel) ajoute les flags/compteurs par type d’atelier
        qs = atelier_tre.AtelierTRE.annotate_candidats_with_atelier_flags(qs)
        return qs
    
    def get_queryset(self):
        return self._scope_qs_to_user_centres(self.base_queryset())

    # ---------- list (log) ----------

    def list(self, request, *args, **kwargs):
        qp_ser = CandidatQueryParamsSerializer(data=request.query_params)
        qp_ser.is_valid(raise_exception=False)
        logger.debug(
            "🧭 qp valid=%s errors=%s cleaned=%s",
            qp_ser.is_valid(), qp_ser.errors, qp_ser.validated_data,
        )

        base_qs = self.get_queryset()
        filtered_qs = self.filter_queryset(base_qs)
        self._log_filters(request, base_qs, filtered_qs)

        return super().list(request, *args, **kwargs)

    # ---------- serializer + context ----------

    def get_serializer_class(self):
        if self.action == "list":
            if self.request.query_params.get("lite") == "1":
                return CandidatLiteSerializer
            return CandidatListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return CandidatCreateUpdateSerializer
        return CandidatSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = getattr(self, "request", None)
        return ctx

    # ---------- create/update : contrôle périmètre formation ----------

    def perform_create(self, serializer):
        instance = serializer.save()
        # Si formation fournie, vérifier périmètre du staff
        self._assert_staff_can_use_formation(getattr(instance, "formation", None))
        try:
            instance.save(user=self.request.user)  # si BaseModel.save(user=...)
        except TypeError:
            pass

    def _cascade_update_prospections_on_formation_change(self, candidat, old_form, new_form):
        """
        Recale toutes les prospections du candidat (owner = candidat.compte_utilisateur)
        - Si new_form est défini  : formation = new_form, centre_id = new_form.centre_id (bulk update)
        - Si new_form est None     : formation = None, centre_id = partenaire.default_centre_id (si dispo) sinon inchangé
                                     (fait en loop, car on dépend du partenaire)
        On ne met à jour QUE les prospections dont la formation était NULL ou == old_form,
        pour éviter d’écraser des choix manuels.
        """
        owner = getattr(candidat, "compte_utilisateur", None)
        if not owner:
            return

        base_qs = Prospection.objects.filter(owner=owner)
        if old_form is not None:
            qs = base_qs.filter(Q(formation__isnull=True) | Q(formation_id=old_form.id))
        else:
            qs = base_qs.filter(formation__isnull=True)

        if not qs.exists():
            return

        if new_form:
            # 🔄 Bulk update quand on peut (plus rapide)
            qs.update(formation=new_form, centre_id=new_form.centre_id)
        else:
            # 🔄 Pas de formation : pour chaque prospection, centre = partenaire.default_centre (si présent)
            for p in qs.select_related("partenaire"):
                p.formation = None
                fallback_centre_id = getattr(getattr(p, "partenaire", None), "default_centre_id", None)
                if fallback_centre_id is not None:
                    p.centre_id = fallback_centre_id
                # sinon on garde le centre existant
                if hasattr(p, "updated_by"):
                    p.updated_by = self.request.user
                    p.save(update_fields=["formation", "centre_id", "updated_by"])
                else:
                    p.save(update_fields=["formation", "centre_id"])

    def perform_update(self, serializer):
        # --- 1) contrôle périmètre ---
        new_formation = serializer.validated_data.get("formation", serializer.instance.formation)
        self._assert_staff_can_use_formation(new_formation)

        # --- 2) détecter changement de formation ---
        old_formation = serializer.instance.formation

        with transaction.atomic():
            instance = serializer.save()
            try:
                instance.save(user=self.request.user)
            except TypeError:
                pass

            # --- 3) cascade sur Prospection si la formation a changé ---
            old_id = getattr(old_formation, "id", None)
            new_id = getattr(instance.formation, "id", None)
            if old_id != new_id:
                self._cascade_update_prospections_on_formation_change(
                    candidat=instance,
                    old_form=old_formation,
                    new_form=instance.formation,
                )

    # ---------- META / EXPORT ----------

    @extend_schema(responses=None)
    @action(
        detail=False,
        methods=["get"],
        url_path="meta",
        url_name="meta",
        renderer_classes=[JSONRenderer],
        permission_classes=[IsAuthenticated],
    )
    def meta(self, request):
        logger.debug("ℹ️ /candidats/meta called")
        data = _build_candidat_meta(request.user)  # ✅ scope staff
        logger.debug("ℹ️ /candidats/meta keys=%s", list(data.keys()))
        return Response(data)

    
    # ---------- Actions Exports----------

    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self.filter_queryset(self.get_queryset())
        logger.debug("📤 export XLSX candidats params=%s rows=%d", self._qp_dict(request), qs.count())

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidats"

        # ==========================================================
        # 🖼️ Logo Rap_App (si dispo)
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre principal
        # ==========================================================
        ws.merge_cells("B1:AF1")
        ws["B1"] = "Export complet des candidats — Rap_App"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:AF2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append([])

        # Ligne de séparation décorative
        sep_row = ws.max_row + 1
        ws.append(["" for _ in range(10)])
        for cell in ws[sep_row]:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
        ws.row_dimensions[sep_row].height = 5

        ws.append([])

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "ID", "Sexe", "Nom de naissance", "Nom d’usage", "Prénom", "Date de naissance",
            "Département de naissance", "Commune de naissance", "Pays de naissance", "Nationalité",
            "NIR", "Âge", "Email", "Téléphone", "Numéro de voie", "Nom de la rue",
            "Complément d’adresse", "Code postal", "Ville", "Statut", "CV", "Type de contrat",
            "Disponibilité", "Entretien réalisé", "Test d’entrée OK", "RQTH", "Permis B",
            "Dernier diplôme préparé", "Diplôme/titre obtenu", "Dernière classe fréquentée",
            "Intitulé diplôme préparé", "Situation avant contrat", "Régime social",
            "Sportif de haut niveau", "Équivalence jeunes", "Extension BOE", "Situation actuelle",
            "Lien représentant", "Nom naissance représentant", "Prénom représentant",
            "Email représentant", "Adresse représentant", "CP représentant", "Ville représentant",
            "Formation", "Num offre", "Centre formation", "Type formation", "Origine sourcing",
            "Date inscription", "Résultat placement", "Contrat signé", "Date placement",
            "Entreprise placement", "Entreprise validée", "Responsable placement", "Vu par (staff)",
            "Nb appairages", "Nb prospections", "Inscrit GESPERS", "Courrier rentrée envoyé",
            "Date rentrée", "Admissible", "OSIA", "Communication ★", "Expérience ★", "CSP ★",
            "Projet création entreprise", "Notes",
        ]
        ws.append(headers)

        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        for i, c in enumerate(qs, start=1):
            ws.append([
                c.id, c.sexe or "", c.nom_naissance or "", c.nom or "", c.prenom or "",
                c.date_naissance.strftime("%d/%m/%Y") if c.date_naissance else "",
                c.departement_naissance or "", c.commune_naissance or "", c.pays_naissance or "",
                c.nationalite or "", c.nir or "", c.age or "", c.email or "", c.telephone or "",
                c.street_number or "", c.street_name or "", c.street_complement or "",
                c.code_postal or "", c.ville or "",
                c.get_statut_display() if hasattr(c, "get_statut_display") else c.statut,
                c.get_cv_statut_display() if hasattr(c, "get_cv_statut_display") else c.cv_statut,
                c.get_type_contrat_display() if hasattr(c, "get_type_contrat_display") else c.type_contrat,
                c.get_disponibilite_display() if hasattr(c, "get_disponibilite_display") else c.disponibilite,
                "Oui" if c.entretien_done else "Non", "Oui" if c.test_is_ok else "Non",
                "Oui" if c.rqth else "Non", "Oui" if c.permis_b else "Non",
                c.dernier_diplome_prepare or "", c.diplome_plus_eleve_obtenu or "",
                c.derniere_classe or "", c.intitule_diplome_prepare or "",
                c.situation_avant_contrat or "", c.regime_social or "",
                "Oui" if c.sportif_haut_niveau else "Non",
                "Oui" if c.equivalence_jeunes else "Non",
                "Oui" if c.extension_boe else "Non", c.situation_actuelle or "",
                c.representant_lien or "", c.representant_nom_naissance or "",
                c.representant_prenom or "", c.representant_email or "",
                c.representant_street_name or "", c.representant_zip_code or "",
                c.representant_city or "",
                getattr(c.formation, "nom", "") if c.formation else "",
                getattr(c.formation, "num_offre", "") if c.formation else "",
                getattr(getattr(c.formation, "centre", None), "nom", "") if c.formation else "",
                getattr(getattr(c.formation, "type_offre", None), "nom", "") if c.formation else "",
                c.origine_sourcing or "",
                c.date_inscription.strftime("%d/%m/%Y") if c.date_inscription else "",
                c.get_resultat_placement_display() if hasattr(c, "get_resultat_placement_display") else c.resultat_placement,
                c.get_contrat_signe_display() if hasattr(c, "get_contrat_signe_display") else c.contrat_signe,
                c.date_placement.strftime("%d/%m/%Y") if c.date_placement else "",
                getattr(c.entreprise_placement, "nom", ""),
                getattr(c.entreprise_validee, "nom", ""),
                getattr(c.responsable_placement, "username", ""),
                getattr(c.vu_par, "username", ""),
                getattr(c, "nb_appairages_calc", 0),
                getattr(c, "nb_prospections_calc", 0),
                "Oui" if c.inscrit_gespers else "Non",
                "Oui" if c.courrier_rentree else "Non",
                c.date_rentree.strftime("%d/%m/%Y") if c.date_rentree else "",
                "Oui" if c.admissible else "Non", c.numero_osia or "",
                c.communication or "", c.experience or "", c.csp or "",
                "Oui" if c.projet_creation_entreprise else "Non",
                (c.notes or "").replace("\n", " "),
            ])

            fill = even_fill if i % 2 == 0 else odd_fill
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10, color="333333")
                cell.alignment = Alignment(vertical="top", wrapText=True)

            ws.row_dimensions[ws.max_row].height = 22

        # ==========================================================
        # 📊 Filtres + gel d’en-tête
        # ==========================================================
        end_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        if end_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        # ==========================================================
        # 📏 Largeurs de colonnes automatiques
        # ==========================================================
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

        # ==========================================================
        # 📈 Pied de page / résumé
        # ==========================================================
        ws.append([])
        ws.append([""])
        ws.append([f"Nombre total de candidats exportés : {qs.count()}"])
        ws[ws.max_row][0].font = Font(name="Calibri", bold=True, color="004C99", size=11)

        ws.oddFooter.center.text = f"© Rap_App — export généré le {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        # ==========================================================
        # 📤 Génération du fichier
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        binary_content = buffer.getvalue()

        filename = f'candidats_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response
