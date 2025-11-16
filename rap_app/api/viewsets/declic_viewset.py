# rap_app/api/viewsets/declic_viewset.py

from django.db.models import Q, Sum
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django.utils.timezone import localdate
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from django.http import HttpResponse
from django.utils import timezone as dj_timezone
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from io import BytesIO
from django.conf import settings

from ...models.centres import Centre
from ...models.declic import Declic, ObjectifDeclic
from ..serializers.declic_serializers import DeclicSerializer
from ..permissions import IsDeclicStaffOrAbove
from ...api.roles import (
    is_admin_like,
    is_staff_or_staffread,
    is_declic_staff,
    is_candidate,
)

# =====================================================================================
# 🔧 HELPERS SCOPE — (MANQUAIENT DANS TON FICHIER) → POUR L'AUTORISATION PAR CENTRES
# =====================================================================================

class ScopeMixin:
    """Mixin réutilisable pour toutes les ViewSets filtrant par centre."""

    def _centre_ids_for_user(self, user):
        """Retourne la liste des centres accessibles selon le rôle."""
        if is_admin_like(user):
            return None  # accès complet

        # déclic_staff = accès restreint à ses centres
        centres = getattr(user, "centres_acces", None) or getattr(user, "centres", None)

        if centres is not None:
            return list(centres.values_list("id", flat=True))

        return []  # par défaut aucun centre

    def _scope_qs_to_user_centres(self, qs):
        """Filtre un queryset sur les centres autorisés."""
        user = self.request.user

        if not user.is_authenticated or is_candidate(user):
            return qs.none()

        centre_ids = self._centre_ids_for_user(user)

        if centre_ids is None:  # superadmin/admin
            return qs

        if len(centre_ids) == 0:
            return qs.none()

        return qs.filter(centre_id__in=centre_ids)


# =====================================================================================
# 📊 DÉCLIC VIEWSET — ATELIERS UNIQUEMENT
# =====================================================================================

@extend_schema_view(
    list=extend_schema(
        summary="Lister tous les ateliers Déclic",
        description="Retourne uniquement les ateliers Déclic (Atelier 1 → 6 + Autre).",
        parameters=[
            OpenApiParameter(name="annee", type=int),
            OpenApiParameter(name="centre", type=int),
            OpenApiParameter(name="departement", type=str),
            OpenApiParameter(name="type_declic", type=str),
            OpenApiParameter(name="date_min", type=str),
            OpenApiParameter(name="date_max", type=str),
            OpenApiParameter(name="search", type=str),
        ],
        responses={200: DeclicSerializer},
    )
)
class DeclicViewSet(ScopeMixin, viewsets.ModelViewSet):
    """📊 Gestion des ateliers Déclic (IC supprimée)."""

    serializer_class = DeclicSerializer
    permission_classes = [IsDeclicStaffOrAbove]
    queryset = Declic.objects.select_related("centre").all()

    # -------------------------------------------------------------------------
    # 🎛️ OPTIONS FILTRES
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):

        annees = (
            Declic.objects.order_by()
            .values_list("date_declic__year", flat=True)
            .distinct()
        )
        annees = sorted([a for a in annees if a], reverse=True)

        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all())

        centres = []
        deps = set()

        for c in centres_qs.order_by("nom"):
            dep = getattr(c, "departement", "") or (c.code_postal[:2] if c.code_postal else "")
            if dep:
                deps.add(dep)

            centres.append({
                "value": c.id,
                "label": c.nom,
                "departement": dep,
                "code_postal": c.code_postal,
            })

        # liste des types atelier uniquement
        types = [{"value": t[0], "label": t[1]} for t in Declic.TypeDeclic.choices]

        return Response({
            "annees": annees,
            "departements": [{"value": d, "label": f"Département {d}"} for d in sorted(deps)],
            "centres": centres,
            "type_declic": types,
        })

    # -------------------------------------------------------------------------
    # 🔍 QUERYSET PRINCIPAL
    # -------------------------------------------------------------------------
    def get_queryset(self):
        qs = Declic.objects.select_related("centre")
        qs = self._scope_qs_to_user_centres(qs)

        p = self.request.query_params
        annee = p.get("annee")
        centre_id = p.get("centre")
        departement = p.get("departement")
        type_declic = p.get("type_declic")
        date_min = p.get("date_min")
        date_max = p.get("date_max")
        search = p.get("search")
        ordering = p.get("ordering")

        if annee:
            qs = qs.filter(date_declic__year=annee)
        if centre_id:
            qs = qs.filter(centre_id=centre_id)
        if type_declic:
            qs = qs.filter(type_declic=type_declic)

        if departement:
            departement = str(departement)
            qs = qs.filter(
                Q(centre__departement__startswith=departement)
                | Q(centre__code_postal__startswith=departement)
            )

        if date_min:
            qs = qs.filter(date_declic__gte=date_min)
        if date_max:
            qs = qs.filter(date_declic__lte=date_max)

        if search:
            qs = qs.filter(
                Q(centre__nom__icontains=search) |
                Q(commentaire__icontains=search)
            )

        return qs.order_by(ordering or "-date_declic", "-id")

    # -------------------------------------------------------------------------
    # 💾 CREATE / UPDATE
    # -------------------------------------------------------------------------
    def perform_create(self, serializer):
        instance = serializer.save()
        instance.save(user=self.request.user)

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.save(user=self.request.user)

    # -------------------------------------------------------------------------
    # 📊 STATISTIQUES
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="stats-centres")
    def stats_centres(self, request):
        annee = int(request.query_params.get("annee", localdate().year))
        data = Declic.accueillis_par_centre(annee)
        return Response(data)

    @action(detail=False, methods=["get"], url_path="stats-departements")
    def stats_departements(self, request):
        annee = int(request.query_params.get("annee", localdate().year))
        data = Declic.accueillis_par_departement(annee)
        return Response(data)

    # -------------------------------------------------------------------------
    # 📤 EXPORT EXCEL ATELIERS UNIQUEMENT
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self.filter_queryset(self.get_queryset())
        annee = int(request.query_params.get("annee", localdate().year))

        wb = Workbook()
        ws = wb.active
        ws.title = "Ateliers Déclic"

        headers = [
            "ID",
            "Atelier",
            "Date",
            "Centre",
            "Inscrits",
            "Présents",
            "Absents",
            "Taux présence (%)",
            "Objectif annuel",
            "Ateliers cumulés",
            "Taux atteinte (%)",
            "Reste à faire",
            "Commentaire",
        ]
        ws.append(headers)

        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for s in qs:
            objectif = s.objectif_annuel

            realise = (
                Declic.objects.filter(
                    centre=s.centre,
                    date_declic__year=s.date_declic.year,
                ).aggregate(total=Sum("nb_presents_declic"))["total"] or 0
            )

            taux = round((realise / objectif) * 100, 1) if objectif else 0

            ws.append([
                s.id,
                s.get_type_declic_display(),
                s.date_declic.strftime("%d/%m/%Y"),
                s.centre.nom if s.centre else "",
                s.nb_inscrits_declic,
                s.nb_presents_declic,
                s.nb_absents_declic,
                s.taux_presence_declic,
                objectif,
                realise,
                taux,
                max(objectif - realise, 0),
                (s.commentaire or "").replace("\n", " "),
            ])

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="declic_ateliers_{annee}.xlsx"'
        return response
