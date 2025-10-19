import logging
from django.db.models import Q
from rest_framework import viewsets, status, filters
from rest_framework.response import Response
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from django.utils import timezone as dj_timezone
from django.conf import settings
from weasyprint import HTML, CSS
from pathlib import Path

from ..paginations import RapAppPagination
from ..permissions import IsStaffOrAbove
from ..roles import is_admin_like, is_staff_like, is_candidate, staff_centre_ids
from ...models.commentaires_appairage import CommentaireAppairage
from ..serializers.commentaires_appairage_serializers import (
    CommentaireAppairageSerializer,
    CommentaireAppairageWriteSerializer,
)
from ...models.logs import LogUtilisateur

logger = logging.getLogger("APPARIAGE_COMMENT")


@extend_schema(tags=["Commentaires Appairages"])
class CommentaireAppairageViewSet(viewsets.ModelViewSet):
    """
    💬 API CRUD + archivage logique pour les commentaires d’appairage.
    Aligné sur la logique des ProspectionComments.
    """

    queryset = CommentaireAppairage.objects.select_related(
        "appairage",
        "appairage__candidat",
        "appairage__partenaire",
        "appairage__formation",
        "appairage__formation__type_offre",
        "created_by",
    ).all()
    serializer_class = CommentaireAppairageSerializer
    permission_classes = [IsStaffOrAbove]
    pagination_class = RapAppPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["appairage", "created_by", "appairage__partenaire", "appairage__formation"]
    search_fields = [
        "body",
        "created_by__username",
        "created_by__email",
        "appairage__candidat__nom",
        "appairage__candidat__prenom",
        "appairage__partenaire__nom",
        "appairage__formation__nom",
    ]
    ordering_fields = ["created_at", "id"]
    ordering = ["-created_at"]

    # ------------------------------------------------------------------
    # 🔍 Queryset dynamique & filtrage par rôle
    # ------------------------------------------------------------------
    def get_queryset(self):
        u = getattr(self.request, "user", None)
        base = super().get_queryset()

        if not getattr(u, "is_authenticated", False):
            return base.none()

        qs = base

        # 🔹 Règle d’accès
        if is_candidate(u):
            qs = qs.filter(appairage__candidat__user_id=u.id)
        elif is_staff_like(u) and not is_admin_like(u):
            centre_ids = staff_centre_ids(u) or []
            if centre_ids:
                qs = qs.filter(appairage__formation__centre_id__in=centre_ids).distinct()

        # 🔹 Filtre "est_archive"
        qp = self.request.query_params
        est_archive = qp.get("est_archive")
        if est_archive is None:
            qs = qs.filter(statut_commentaire="actif")
        else:
            val = est_archive.lower()
            if val in ("1", "true", "yes", "oui"):
                qs = qs.filter(statut_commentaire="archive")
            elif val in ("0", "false", "no", "non"):
                qs = qs.filter(statut_commentaire="actif")

        # 🔹 Filtres additionnels
        partenaire_nom = (qp.get("partenaire_nom") or "").strip()
        candidat_nom = (qp.get("candidat_nom") or "").strip()
        formation_nom = (qp.get("formation_nom") or "").strip()
        if partenaire_nom:
            qs = qs.filter(appairage__partenaire__nom__icontains=partenaire_nom)
        if candidat_nom:
            qs = qs.filter(appairage__candidat__nom__icontains=candidat_nom)
        if formation_nom:
            qs = qs.filter(appairage__formation__nom__icontains=formation_nom)

        return qs.order_by("-created_at", "-id").distinct()

    # ------------------------------------------------------------------
    # 💾 Sérialiseurs dynamiques
    # ------------------------------------------------------------------
    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return CommentaireAppairageWriteSerializer
        return CommentaireAppairageSerializer

    # ------------------------------------------------------------------
    # ➕ CREATE / ✏️ UPDATE / 🗑️ DELETE
    # ------------------------------------------------------------------
    def perform_create(self, serializer):
        commentaire = serializer.save(created_by=self.request.user)
        LogUtilisateur.log_action(
            instance=commentaire,
            action=LogUtilisateur.ACTION_CREATE,
            user=self.request.user,
            details=f"Création d’un commentaire pour l’appairage #{commentaire.appairage_id}",
        )

    def perform_update(self, serializer):
        commentaire = serializer.save(updated_by=self.request.user)
        LogUtilisateur.log_action(
            instance=commentaire,
            action=LogUtilisateur.ACTION_UPDATE,
            user=self.request.user,
            details=f"Mise à jour du commentaire d’appairage #{commentaire.pk}",
        )

    def perform_destroy(self, instance):
        pk = instance.pk
        instance.delete()
        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_DELETE,
            user=self.request.user,
            details=f"Suppression du commentaire d’appairage #{pk}",
        )

    # ------------------------------------------------------------------
    # 🔒 ARCHIVER / DÉSARCHIVER
    # ------------------------------------------------------------------
    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        """🗃️ Archive logiquement un commentaire d’appairage."""
        comment = self.get_object()
        if comment.est_archive:
            return Response({"detail": "Déjà archivé."}, status=status.HTTP_200_OK)
        comment.archiver(save=True)
        logger.info("CommentaireAppairage #%s archivé par %s", comment.pk, request.user)
        return Response({"detail": "Commentaire archivé."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """♻️ Restaure un commentaire archivé."""
        comment = self.get_object()
        if not comment.est_archive:
            return Response({"detail": "Déjà actif."}, status=status.HTTP_200_OK)
        comment.desarchiver(save=True)
        logger.info("CommentaireAppairage #%s désarchivé par %s", comment.pk, request.user)
        return Response({"detail": "Commentaire désarchivé."}, status=status.HTTP_200_OK)

    # ------------------------------------------------------------------
    # 📊 EXPORT XLSX — Commentaires d’appairage
    # ------------------------------------------------------------------
    @extend_schema(summary="Exporter les commentaires d’appairage au format XLSX")
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self.filter_queryset(self.get_queryset())
        wb = Workbook()
        ws = wb.active
        ws.title = "Commentaires Appairage"

        # Logo Rap_App
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height, img.width = 45, 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        # Titre + date
        ws.merge_cells("B1:H1")
        ws["B1"] = "Commentaires d’appairage — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="0077CC")
        ws["B1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("B2:H2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center")

        ws.append([])

        headers = [
            "ID",
            "Statut commentaire",
            "Appairage",
            "Candidat",
            "Partenaire",
            "Formation",
            "Auteur",
            "Commentaire",
            "Créé le",
        ]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9F2FF")
            cell.alignment = Alignment(horizontal="center")

        for c in qs:
            statut_display = dict(c.STATUT_CHOICES).get(c.statut_commentaire, c.statut_commentaire)
            statut_color = "C8E6C9" if c.statut_commentaire == "actif" else "E0E0E0"

            ws.append([
                c.id,
                statut_display,
                getattr(c.appairage, "id", "—"),
                getattr(getattr(c.appairage, "candidat", None), "nom_complet", "—"),
                getattr(getattr(c.appairage, "partenaire", None), "nom", "—"),
                getattr(getattr(c.appairage, "formation", None), "nom", "—"),
                getattr(c.created_by, "username", "—"),
                c.body or "",
                c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "—",
            ])
            statut_cell = ws[f"B{ws.max_row}"]
            statut_cell.fill = PatternFill("solid", fgColor=statut_color)
            statut_cell.font = Font(bold=True)

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == "H":
                ws.column_dimensions[col_letter].width = 80
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            else:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        binary = buffer.getvalue()
        filename = f'appairage_commentaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary)
        return response

    # ------------------------------------------------------------------
    # 📄 EXPORT PDF — Commentaires d’appairage
    # ------------------------------------------------------------------
    @extend_schema(summary="Exporter les commentaires d’appairage au format PDF")
    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        qs = self.filter_queryset(self.get_queryset())

        try:
            logo_url = request.build_absolute_uri(static("images/logo.png"))
        except Exception:
            logo_path = settings.BASE_DIR / "rap_app/static/images/logo.png"
            logo_url = f"file://{logo_path}"

        # 🧩 On enrichit les données avant de les passer au template
        commentaires_data = []
        for c in qs:
            commentaires_data.append({
                "id": c.id,
                "statut_commentaire": c.statut_commentaire,
                "body": c.body or "",
                "created_by": getattr(c.created_by, "username", "—"),
                "created_at": c.created_at,
                "activite": getattr(c, "activite", ""),  # ✅ ajouté ici
                "appairage": {
                    "id": getattr(c.appairage, "id", None),
                    "statut": getattr(c.appairage, "statut", ""),
                    "get_statut_display": getattr(c.appairage, "get_statut_display", lambda: "")(),
                    "candidat": {
                        "nom_complet": getattr(getattr(c.appairage, "candidat", None), "nom_complet", "—"),
                    },
                    "partenaire": {
                        "nom": getattr(getattr(c.appairage, "partenaire", None), "nom", "—"),
                    },
                    "formation": {
                        "nom": getattr(getattr(c.appairage, "formation", None), "nom", "—"),
                        "num_offre": getattr(getattr(c.appairage, "formation", None), "num_offre", ""),
                        "type_offre": {
                            "nom": getattr(
                                getattr(getattr(c.appairage, "formation", None), "type_offre", None),
                                "nom",
                                "—",
                            ),
                        },
                        "start_date": getattr(
                            getattr(c.appairage, "formation", None), "start_date", None
                        ),
                        "end_date": getattr(
                            getattr(c.appairage, "formation", None), "end_date", None
                        ),
                    },
                },
            })

        # 🔹 Contexte final
        context = {
            "commentaires": commentaires_data,
            "now": dj_timezone.now(),
            "logo_url": logo_url,
            "user": request.user,
        }

        html_string = render_to_string("exports/appairage_commentaires_pdf.html", context)
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

        filename = f'commentaires_appairage_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(pdf)
        return response
